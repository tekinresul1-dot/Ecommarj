from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from core.permissions import IsSubscribed
from django.utils import timezone
from django.db import models as django_models
from decimal import Decimal

from django.db.models import Prefetch
from core.models import (
    Organization, MarketplaceAccount, Product, ProductVariant,
    Order, OrderItem, FinancialTransactionType
)
from core.utils.mixins import TenantQuerySetMixin
from core.services.profit_calculator import ProfitCalculator
from core.tasks import sync_all_trendyol_data_task

MONTHS_TR = {
    1: "Oca", 2: "Şub", 3: "Mar", 4: "Nis", 5: "May", 6: "Haz",
    7: "Tem", 8: "Ağu", 9: "Eyl", 10: "Eki", 11: "Kas", 12: "Ara"
}

# ---------------------------------------------------------------------------
# Sipariş statüsü sabitleri — tüm view'larda tutarlı kullanım için
# ---------------------------------------------------------------------------
# Satış olarak sayılan: sadece teslim onaylanmış siparişler
# Shipped (kargoda) henüz teslim kesinleşmedi → Trendyol da saymıyor
SALE_STATUSES = ["Delivered"]
# Kargo hareketi olan iade/teslim edilemedi — kargo zararı oluşur
RETURN_STATUSES = ["Returned", "UnDelivered"]
# Sevkiyat yapılmamış iptal/tedarik edilemedi — kargo zararı yok
CANCEL_STATUSES = ["Cancelled", "UnSupplied"]
# Satış sayılmayan tüm statüler
NON_SALE_STATUSES = RETURN_STATUSES + CANCEL_STATUSES

def format_date_tr(dt):
    if not dt:
        return ""
    from datetime import datetime
    # Trendyol timestamps are stored as Istanbul local time (no conversion needed)
    if hasattr(dt, 'replace'):
        dt = dt.replace(tzinfo=None)
    return f"{dt.day:02d} {MONTHS_TR[dt.month]} {dt.year} - {dt.strftime('%H:%M') if isinstance(dt, datetime) else '00:00'}"

def format_date_short_tr(dt):
    if not dt:
        return ""
    from datetime import datetime, date
    if isinstance(dt, datetime):
        dt = dt.replace(tzinfo=None)
    return f"{dt.day:02d} {MONTHS_TR[dt.month]}"
    
class TriggerSyncView(APIView):
    """
    Dashboard'dan manuel "Trendyol Senkronize Et" butonuna basıldığında tetiklenir.
    Kullanıcının Trendyol hesaplarını bulup Celery task'ine gönderir.
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def post(self, request):
        user = request.user
        # Auto-create profile and org if missing (useful for superusers or first time testing)
        from core.models import UserProfile, Organization
        profile, _ = UserProfile.objects.get_or_create(user=user)
        org = profile.organization
        if not org:
            org, _ = Organization.objects.get_or_create(name=f"{user.username} Organizasyonu")
            profile.organization = org
            profile.save()
            
        accounts = MarketplaceAccount.objects.filter(organization=org, channel=MarketplaceAccount.Channel.TRENDYOL, is_active=True)
        
        if not accounts.exists():
            return Response({"error": "Aktif bir Trendyol API hesabı bulunamadı."}, status=404)
            
        task_ids = []
        for acc in accounts:
            task = sync_all_trendyol_data_task.delay(str(acc.id))
            task_ids.append(task.id)

        return Response({
            "message": f"{accounts.count()} hesap için senkronizasyon başlatıldı.",
            "task_ids": task_ids,
        })


class ProductStockSyncView(APIView):
    """
    POST /api/products/sync-stock/
    Sadece ürün stok bilgilerini Trendyol'dan çeker ve günceller (sipariş/hakediş senkronizasyonu yapmaz).
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def post(self, request):
        from core.models import UserProfile
        from core.services.sync_service import TrendyolSyncService
        from core.utils.encryption import decrypt_value

        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        org = profile.organization
        if not org:
            return Response({"error": "Organizasyon bulunamadı."}, status=404)

        accounts = MarketplaceAccount.objects.filter(
            organization=org,
            channel=MarketplaceAccount.Channel.TRENDYOL,
            is_active=True
        )
        if not accounts.exists():
            return Response({"error": "Aktif bir Trendyol API hesabı bulunamadı. Lütfen Ayarlar > Trendyol bölümünden API kimlik bilgilerinizi ekleyin."}, status=404)

        synced_count = 0
        for acc in accounts:
            try:
                service = TrendyolSyncService(acc)
                service.sync_products()
                synced_count += 1
            except ValueError as e:
                return Response({"error": str(e)}, status=400)
            except Exception as e:
                return Response({"error": f"Stok senkronizasyonu sırasında hata oluştu: {str(e)}"}, status=500)

        return Response({"message": f"Stok bilgileri güncellendi. {synced_count} Trendyol hesabı senkronize edildi."})


class DashboardOverviewView(APIView):
    """
    GET /api/dashboard/overview
    Filtrelere göre Sipariş, Ürün ve Finansal verileri toplayıp gerçek KPI döndürür.
    
    Toplam Ciro = Sipariş kayıtlarındaki sale_price_net toplamı (Trendyol'un "amount" alanı — indirimli satış fiyatı)
    Maliyetlendirilen Ciro = Sadece maliyeti tanımlı ürünlerin cirosu
    Kâr Tutarı = Maliyetlendirilen ciro üzerinden ProfitCalculator ile hesaplanan kâr
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request):
        user = request.user
        from core.models import UserProfile, Organization
        from datetime import datetime as dt_cls, time as dt_time, timezone as dt_tz
        from collections import defaultdict
        
        profile, _ = UserProfile.objects.get_or_create(user=user)
        org = profile.organization
        if not org:
            org, _ = Organization.objects.get_or_create(name=f"{user.username} Organizasyonu")
            profile.organization = org
            profile.save()

        # Filtreleri al
        channel = request.query_params.get("channel", "trendyol")
        countries = request.query_params.get("countries", "")
        min_date_str = request.query_params.get("min_date")
        max_date_str = request.query_params.get("max_date")

        # Organizasyonun siparişlerini çek
        # Trendyol panelinde domestic ve micro_export birlikte gösteriliyor
        if channel == "trendyol":
            orders_qs = Order.objects.select_related("marketplace_account").filter(
                organization=org, channel__in=["trendyol", "micro_export"]
            )
        else:
            orders_qs = Order.objects.select_related("marketplace_account").filter(
                organization=org, channel=channel
            )

        if countries:
            country_list = [c.strip() for c in countries.split(",")]
            orders_qs = orders_qs.filter(country_code__in=country_list)

        if min_date_str and max_date_str:
            try:
                min_date = dt_cls.strptime(min_date_str, "%Y-%m-%d").replace(tzinfo=dt_tz.utc)
                max_date = dt_cls.combine(
                    dt_cls.strptime(max_date_str, "%Y-%m-%d"), dt_time.max
                ).replace(tzinfo=dt_tz.utc)
                orders_qs = orders_qs.filter(order_date__gte=min_date, order_date__lte=max_date)
            except ValueError:
                pass

        # ---------------------------------------------------------------
        # 1. TOPLAM CİRO = Direkt sipariş kayıtlarından (Trendyol gibi)
        #    - Tüm siparişlerin "sale_price_net" (indirimli satış tutarı) toplamı
        #    - İptal ve iade düşülmüş hali (aktif siparişler)
        # ---------------------------------------------------------------
        total_orders = orders_qs.count()

        # Sadece gerçekleşmiş satışlar: Delivered + Shipped
        # UnDelivered / Cancelled / Returned / UnSupplied / Picking / Created satış değil
        active_orders_qs = orders_qs.filter(status__in=SALE_STATUSES)
        cancelled_orders_qs = orders_qs.filter(status__in=NON_SALE_STATUSES)

        active_order_count = active_orders_qs.count()
        cancelled_count = cancelled_orders_qs.count()

        # Toplam Ciro: Aktif sipariş kalemlerinin sale_price_net toplamı
        # Ürün maliyet uyarısı için sayım (filtre öncesi)
        from core.models import ProductVariant, Product as ProductModel
        _total_org_products = ProductModel.objects.filter(organization=org).count()
        _products_with_cost = ProductVariant.objects.filter(
            product__organization=org,
            cost_price__isnull=False,
            cost_price__gt=0
        ).values("product").distinct().count()
        _products_without_cost = _total_org_products - _products_with_cost

        # Sadece maliyeti girilmiş ürünlerin sipariş kalemlerini hesapla
        active_items_qs = OrderItem.objects.filter(
            order__in=active_orders_qs,
            product_variant__cost_price__isnull=False,
            product_variant__cost_price__gt=0
        )
        
        revenue_agg = active_items_qs.aggregate(
            total_gross=django_models.Sum("sale_price_gross"),
            total_net=django_models.Sum("sale_price_net"),
            total_discount=django_models.Sum("discount"),
            total_items=django_models.Count("id"),
            total_quantity=django_models.Sum("quantity"),
        )
        
        toplam_ciro = revenue_agg["total_net"] or Decimal("0.00")
        total_gross = revenue_agg["total_gross"] or Decimal("0.00")
        total_discount = revenue_agg["total_discount"] or Decimal("0.00")
        total_items_sold = revenue_agg["total_quantity"] or 0
        
        # ---------------------------------------------------------------
        # 2. KARLILIK HESABI = ProfitCalculator ile (sadece maliyeti olan ürünler)
        # ---------------------------------------------------------------
        total_costs = Decimal("0.00")
        total_profit = Decimal("0.00")
        total_costed_revenue = Decimal("0.00")
        
        breakdown = defaultdict(Decimal)
        for tx_type in FinancialTransactionType:
            breakdown[tx_type.value] = Decimal("0.00")
        country_profits = {}
        
        total_cargo_cost = Decimal("0.00")
        total_commission_cost = Decimal("0.00")
        
        funnel_gross = Decimal("0.00")
        funnel_cargo = Decimal("0.00")
        funnel_marketplace_fees = Decimal("0.00")
        funnel_taxes = Decimal("0.00")
        funnel_net = Decimal("0.00")

        # ProfitCalculator ile her satırı tek tek hesapla
        for item in active_items_qs.select_related("order").prefetch_related("transactions"):
            res = ProfitCalculator.calculate_for_order_item(item)
            
            item_revenue = res["gross_revenue"]  # ProfitCalculator'ın döndüğü = sale_price_net
            costs = res["total_costs"]
            profit = res["net_profit"]
            
            total_costs += costs
            total_profit += profit
            
            item_product_cost = res["breakdown"].get(FinancialTransactionType.PRODUCT_COST.value, Decimal("0.00"))
            if item_product_cost > Decimal("0.00"):
                total_costed_revenue += item.sale_price_net  # Gerçek satış fiyatı
            
            funnel_gross += item.sale_price_net  # Gerçek satış fiyatı
            
            item_cargo = Decimal("0.00")
            item_mp_fees = Decimal("0.00")
            item_taxes = Decimal("0.00")
            
            for k, v in res["breakdown"].items():
                breakdown[k] += v
                if k == FinancialTransactionType.SHIPPING_FEE.value:
                    item_cargo += v
                    total_cargo_cost += v
                elif k == FinancialTransactionType.COMMISSION.value:
                    item_mp_fees += v
                    total_commission_cost += v
                elif k in [FinancialTransactionType.SERVICE_FEE.value, FinancialTransactionType.PENALTY.value]:
                    item_mp_fees += v
                elif k in [FinancialTransactionType.VAT_OUTPUT.value, FinancialTransactionType.WITHHOLDING.value]:
                    item_taxes += v
            
            funnel_cargo += item_cargo
            funnel_marketplace_fees += item_mp_fees
            funnel_taxes += item_taxes
            
            c_code = item.order.country_code
            country_profits[c_code] = country_profits.get(c_code, Decimal("0.00")) + profit

        funnel_net = total_profit

        # Kâr Marjları
        profit_margin = Decimal("0.00")
        if toplam_ciro > Decimal("0.00"):
            profit_margin = (total_profit / toplam_ciro) * Decimal("100.00")
            
        profit_on_cost_ratio = Decimal("0.00")
        total_product_cost = breakdown.get(FinancialTransactionType.PRODUCT_COST.value, Decimal("0.00"))
        if total_product_cost > Decimal("0.00"):
            profit_on_cost_ratio = (total_profit / total_product_cost) * Decimal("100.00")

        # Kayıp Kaçak
        lost_and_leakage = (
            breakdown.get(FinancialTransactionType.PENALTY.value, Decimal("0")) +
            breakdown.get(FinancialTransactionType.RETURN_LOSS.value, Decimal("0")) +
            breakdown.get(FinancialTransactionType.EARLY_PAYMENT.value, Decimal("0"))
        )
        
        # --- Metrik Listeleri ---
        
        order_metrics = {
            "order_count": active_order_count,
            "total_order_count": total_orders,
            "cancelled_count": cancelled_count,
            "avg_revenue_per_order": str(round(toplam_ciro / active_order_count, 2)) if active_order_count > 0 else "0.00",
            "avg_profit_per_order": str(round(total_profit / active_order_count, 2)) if active_order_count > 0 else "0.00",
            "avg_cargo_per_order": str(round(total_cargo_cost / active_order_count, 2)) if active_order_count > 0 else "0.00",
        }
        
        product_metrics = {
            "items_sold": total_items_sold,
            "avg_revenue_per_item": str(round(toplam_ciro / total_items_sold, 2)) if total_items_sold > 0 else "0.00",
            "avg_profit_per_item": str(round(total_profit / total_items_sold, 2)) if total_items_sold > 0 else "0.00",
            "avg_cargo_per_item": str(round(total_cargo_cost / total_items_sold, 2)) if total_items_sold > 0 else "0.00",
            "avg_commission_rate": str(round((total_commission_cost / toplam_ciro)*100, 2)) if toplam_ciro > 0 else "0.00",
            "avg_discount_rate": str(round((total_discount / total_gross)*100, 2)) if total_gross > 0 else "0.00",
        }
        
        # İade Metrikleri
        return_loss_val = breakdown.get(FinancialTransactionType.RETURN_LOSS.value, Decimal("0.00"))
        # İade oranı: kargo hareketi olan Returned+UnDelivered + iptal edilen Cancelled+UnSupplied
        returned_orders_count = orders_qs.filter(status__in=NON_SALE_STATUSES).count()
        real_return_rate = Decimal("0.00")
        if total_orders > 0:
            real_return_rate = round(Decimal(returned_orders_count) / Decimal(total_orders) * Decimal("100"), 2)

        # Kargo zararı: sadece kargoya verilip dönen siparişler (Cancelled'da kargo hareketi yok)
        returned_items = OrderItem.objects.filter(order__in=orders_qs.filter(status__in=RETURN_STATUSES))
        total_return_cargo_loss = Decimal("0.00")
        for r_item in returned_items:
            r_info = ProfitCalculator.calculate_for_order_item(r_item)
            total_return_cargo_loss += r_info["breakdown"].get(FinancialTransactionType.SHIPPING_FEE.value, Decimal("0.00"))
        
        return_metrics = {
            "return_rate": str(real_return_rate),
            "total_return_cost": str(return_loss_val + total_return_cargo_loss),
            "return_cargo_loss": str(total_return_cargo_loss),
            "overseas_operation_fee": "0.00"
        }
        
        # Reklam Metrikleri
        ads_val = breakdown.get(FinancialTransactionType.ADS_COST.value, Decimal("0.00"))
        ads_metrics = {
            "total_ads_cost": str(ads_val),
            "influencer_cut": "0.00",
            "ads_profit_index": str(round((ads_val / total_profit)*100, 2)) if total_profit > 0 else "0.00",
            "ads_revenue_index": str(round((ads_val / toplam_ciro)*100, 2)) if toplam_ciro > 0 else "0.00",
        }
        
        # Net Kâr Performansı Hunisi
        net_profit_funnel = [
            {"name": "Ciro", "value": str(funnel_gross)},
            {"name": "Kargo Düşülmüş", "value": str(funnel_gross - funnel_cargo)},
            {"name": "Pazaryeri Masrafı Düşülmüş", "value": str(funnel_gross - funnel_cargo - funnel_marketplace_fees)},
            {"name": "Vergi Düşülmüş", "value": str(funnel_gross - funnel_cargo - funnel_marketplace_fees - funnel_taxes)},
            {"name": "Kâr", "value": str(funnel_net)},
        ]
        
        # Kâr Performansı History (Area Chart)
        from datetime import timedelta
        from django.db.models.functions import TruncDate
        from django.db.models import Sum
        
        history = []
        day_range = 30
        start_day = timezone.now() - timedelta(days=day_range)
        daily_orders = active_orders_qs.filter(order_date__gte=start_day).annotate(day=TruncDate('order_date')).values('day').annotate(
            day_revenue=Sum('items__sale_price_net')
        ).order_by('day')
        
        daily_profit_map = {}
        total_daily_revenue = sum(d['day_revenue'] or 0 for d in daily_orders)
        for d in daily_orders:
            day_str = format_date_short_tr(d['day']) if d['day'] else "?"
            day_revenue = d['day_revenue'] or Decimal("0")
            if total_daily_revenue > 0:
                day_profit = total_profit * (day_revenue / Decimal(str(total_daily_revenue)))
            else:
                day_profit = Decimal("0")
            daily_profit_map[day_str] = round(day_profit, 2)
        
        for i in range(13, -1, -1):
            day = format_date_short_tr(timezone.now() - timedelta(days=i))
            history.append({
                "date": day,
                "profit": str(daily_profit_map.get(day, Decimal("0.00")))
            })

        # Kritik Stok Uyarıları
        from core.models import Product
        org_products = Product.objects.filter(organization=org, is_active=True)
        low_stock_list = []
        for p in org_products:
            if p.is_low_stock:
                low_stock_list.append({
                    "id": p.id,
                    "title": p.title,
                    "barcode": p.barcode,
                    "current_stock": p.current_stock,
                    "initial_stock": p.initial_stock,
                    "image_url": p.image_url
                })
        
        low_stock_list = sorted(low_stock_list, key=lambda x: x["current_stock"])[:5]

        return Response({
            "kpis": {
                "total_orders": active_order_count,
                "gross_revenue": str(total_gross),
                "toplam_ciro": str(toplam_ciro),          # Toplam Ciro = SUM(sale_price_net) from order items
                "costed_revenue": str(total_costed_revenue),  # Maliyetlendirilen Ciro
                "net_revenue": str(toplam_ciro),           # Keep backward compat
                "total_costs": str(total_costs),
                "net_profit": str(total_profit),
                "profit_margin": str(round(profit_margin, 2)),
                "profit_on_cost_ratio": str(round(profit_on_cost_ratio, 2)),
                "lost_and_leakage": str(lost_and_leakage),
            },
            "profit_breakdown": {k: str(v) for k, v in breakdown.items()},
            "country_profit": [{"country": k, "profit": str(v)} for k, v in country_profits.items()],
            "order_metrics": order_metrics,
            "product_metrics": product_metrics,
            "return_metrics": return_metrics,
            "ads_metrics": ads_metrics,
            "net_profit_funnel": net_profit_funnel,
            "profit_performance_history": history,
            "low_stock_alerts": low_stock_list,
            "products_without_cost_warning": {
                "count": _products_without_cost,
                "total": _total_org_products,
            },
        })


class ProductCostStatusView(APIView):
    """
    GET /api/user/product-cost-status/
    Kullanıcının ürünlerin maliyet durumunu döndürür (popup için).
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request):
        from core.models import Product, ProductVariant, MarketplaceAccount, UserProfile, Organization

        try:
            profile = request.user.profile
            org = profile.organization
        except Exception:
            return Response({"has_costs": False, "total_products": 0, "products_without_cost": 0, "products_with_cost": 0}, status=200)

        if not org:
            return Response({"has_costs": False, "total_products": 0, "products_without_cost": 0, "products_with_cost": 0}, status=200)

        has_marketplace = MarketplaceAccount.objects.filter(
            organization=org, is_active=True
        ).exists()

        total_products = Product.objects.filter(organization=org).count()
        products_with_cost = ProductVariant.objects.filter(
            product__organization=org,
            cost_price__isnull=False,
            cost_price__gt=0
        ).values("product").distinct().count()
        products_without_cost = total_products - products_with_cost

        return Response({
            "has_costs": products_with_cost > 0,
            "has_marketplace_account": has_marketplace,
            "total_products": total_products,
            "products_with_cost": products_with_cost,
            "products_without_cost": products_without_cost,
        })


class MockReportsView(APIView):
    """
    Diğer listeleme sayfaları için geçici endpoint (UI'ın çalışması adına boş veya mock liste döner)
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request, report_type):
        return Response({
            "report_type": report_type,
            "data": []
        })


class TrendyolTestConnectionView(APIView):
    """
    POST /api/integrations/trendyol/test-connection/
    Trendyol bağlantısını test eder (1 ürün çekmeye çalışır).
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def post(self, request):
        import requests as http_requests
        import logging
        logger = logging.getLogger(__name__)

        api_key = request.data.get("api_key", "").strip()
        api_secret = request.data.get("api_secret", "").strip()
        supplier_id = request.data.get("supplier_id", "").strip()

        if not all([api_key, api_secret, supplier_id]):
            return Response({"ok": False, "message": "Eksik bilgi: api_key, api_secret veya supplier_id gerekli."}, status=400)

        # CRITICAL: Correct base URL per official Trendyol docs:
        # apigw.trendyol.com/integration/ (NOT api.trendyol.com/sapigw/ which is Cloudflare-blocked)
        url = f"https://apigw.trendyol.com/integration/product/sellers/{supplier_id}/products"
        params = {"approved": "true", "page": 0, "size": 1}
        headers = {
            "User-Agent": f"{supplier_id} - SelfIntegration",
            "Accept": "application/json",
        }

        # Debug logging (secrets redacted)
        logger.info(f"[TestConnection] URL: {url}")
        logger.info(f"[TestConnection] Params: {params}")
        logger.info(f"[TestConnection] User-Agent: {headers['User-Agent']}")
        logger.info(f"[TestConnection] Auth: ({api_key[:4]}...***)")
        
        # Detect server public IP for diagnostics
        server_ip = "bilinmiyor"
        try:
            ip_resp = http_requests.get("https://api.ipify.org", timeout=5)
            if ip_resp.ok:
                server_ip = ip_resp.text.strip()
        except Exception:
            pass
        
        try:
            res = http_requests.get(url, auth=(api_key, api_secret), params=params, headers=headers, timeout=30)
            
            # Debug: Log response info
            request_id = res.headers.get("x-request-id", "N/A")
            logger.info(f"[TestConnection] Status: {res.status_code}, x-request-id: {request_id}")
            logger.info(f"[TestConnection] Response body: {res.text}")
            
            if not res.ok:
                text = res.text
                status = res.status_code
                message = text[:300]
                
                # HTML check (Cloudflare block)
                if '<html' in text.lower() or 'cloudflare' in text.lower():
                    message = (
                        f"Cloudflare tarafından engellendiniz (HTTP {status}). "
                        f"Sunucu IP adresiniz: {server_ip} — "
                        "Bu IP'nin Trendyol Satıcı Paneli'nde API whitelist'ine eklendiğinden emin olun. "
                        "Ekleme sonrası yayılması 10-30 dakika sürebilir."
                    )
                elif status == 401:
                    message = "API Key veya API Secret hatalı. Lütfen kontrol edip tekrar deneyin."
                elif status == 403:
                    message = f"Erişim reddedildi (403). Sunucu IP: {server_ip}. Trendyol whitelist'ini kontrol edin."
                elif status == 429:
                    message = "Çok fazla istek gönderildi. Lütfen biraz bekleyip tekrar deneyin."
                else:
                    try:
                        data = res.json()
                        message = data.get("message", text[:300])
                    except ValueError:
                        pass
                
                return Response({
                    "ok": False,
                    "code": status,
                    "message": message,
                    "request_id": request_id,
                    "server_ip": server_ip,
                }, status=400)
                
            data = res.json()
            return Response({
                "ok": True,
                "sample_product_count_hint": data.get("totalElements", 0),
                "request_id": request_id,
            })
            
        except http_requests.RequestException as e:
            logger.error(f"[TestConnection] Network error: {e}")
            return Response({"ok": False, "message": f"Bağlantı hatası: {str(e)}", "code": 500}, status=500)


class TrendyolSaveCredentialsView(APIView):
    """
    POST /api/integrations/trendyol/save-credentials/
    Kullanıcının Trendyol API Key, Secret ve Satıcı ID'sini veritabanına şifreli kaydeder.
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request):
        user = request.user
        from core.models import UserProfile, Organization
        from core.utils.encryption import decrypt_value
        profile, _ = UserProfile.objects.get_or_create(user=user)
        org = profile.organization
        if not org:
            org, _ = Organization.objects.get_or_create(name=f"{user.username} Organizasyonu")
            profile.organization = org
            profile.save()
            
        account = MarketplaceAccount.objects.filter(
            organization=org, 
            channel=MarketplaceAccount.Channel.TRENDYOL
        ).first()

        if not account:
            return Response({
                "api_key": "",
                "supplier_id": ""
            })

        return Response({
            "api_key": decrypt_value(account.api_key),
            "supplier_id": account.seller_id
        })

    def post(self, request):
        user = request.user
        from core.models import UserProfile, Organization
        from core.utils.encryption import encrypt_value
        profile, _ = UserProfile.objects.get_or_create(user=user)
        org = profile.organization
        if not org:
            org, _ = Organization.objects.get_or_create(name=f"{user.username} Organizasyonu")
            profile.organization = org
            profile.save()

        api_key = request.data.get("api_key", "").strip()
        api_secret = request.data.get("api_secret", "").strip()
        supplier_id = request.data.get("supplier_id", "").strip()

        if not supplier_id:
            return Response({"error": "Satıcı ID (Supplier ID) zorunludur."}, status=400)

        # Mağaza limiti kontrolü — sadece yeni mağaza ekleniyorsa kontrol et
        from core.services.subscription_service import check_store_limit
        existing_account = MarketplaceAccount.objects.filter(
            organization=org, channel=MarketplaceAccount.Channel.TRENDYOL
        ).exists()
        if not existing_account:
            allowed, msg = check_store_limit(request.user)
            if not allowed:
                return Response({"error": msg, "upgrade_url": "/subscription"}, status=403)

        account, created = MarketplaceAccount.objects.update_or_create(
            organization=org,
            channel=MarketplaceAccount.Channel.TRENDYOL,
            defaults={
                "store_name": f"Trendyol Store - {supplier_id}",
                "seller_id": supplier_id,
                "api_key": encrypt_value(api_key),
                "api_secret": encrypt_value(api_secret),
                "is_active": True
            }
        )

        auto_sync = request.data.get("auto_sync", True)
        task_id = None
        if auto_sync and account.api_key and account.api_secret:
            task = sync_all_trendyol_data_task.delay(str(account.id))
            task_id = task.id

        return Response({
            "message": "Trendyol API bilgileri başarıyla kaydedildi.",
            "sync_started": auto_sync,
            "task_id": task_id,
        })


class ProductListView(APIView):
    """
    GET /api/products/
    Veritabanına (Trendyol üzerinden) senkronize edilmiş ürünlerin tablo olarak dökümünü sağlar.
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request):
        user = request.user
        from core.models import UserProfile, Organization
        profile, _ = UserProfile.objects.get_or_create(user=user)
        org = profile.organization
        if not org:
            org, _ = Organization.objects.get_or_create(name=f"{user.username} Organizasyonu")
            profile.organization = org
            profile.save()

        from core.models import Product
        from django.db.models import Q
        from rest_framework.pagination import PageNumberPagination
        from django.utils import timezone
        import datetime
        
        two_months_ago = timezone.now() - datetime.timedelta(days=60)
        
        from django.db.models import Max
        
        # Paginate by unique marketplace_skus (Model Codes)
        skus_qs = Product.objects.filter(
            organization=org,
            is_active=True
        ).values('marketplace_sku').annotate(
            latest_created_at=Max('trendyol_created_at')
        ).order_by('-latest_created_at')

        search = request.GET.get('search', '').strip()
        if search:
            skus_qs = skus_qs.filter(
                Q(title__icontains=search) |
                Q(barcode__icontains=search) |
                Q(marketplace_sku__icontains=search) |
                Q(variants__barcode__icontains=search) |
                Q(variants__title__icontains=search)
            ).distinct()
            
        paginator = PageNumberPagination()
        paginator.page_size = 50
        paginator.page_size_query_param = 'page_size'
        paginator.max_page_size = 100
        
        paginated_skus = paginator.paginate_queryset(skus_qs, request)
        current_skus = [item['marketplace_sku'] for item in paginated_skus]
        
        # Fetch full data for these SKUs
        products = Product.objects.filter(
            organization=org,
            marketplace_sku__in=current_skus,
            is_active=True
        ).prefetch_related('variants').order_by(
            django_models.F('trendyol_created_at').desc(nulls_last=True)
        )
        
        # Simple serialization
        data = []
        for p in products:
            product_data = {
                "id": p.id,
                "title": p.title,
                "barcode": p.barcode,
                "marketplace_sku": p.marketplace_sku,
                "trendyol_content_id": p.trendyol_content_id,
                "currency": p.currency,
                "sale_price": str(p.sale_price),
                "vat_rate": str(p.vat_rate),
                "commission_rate": str(p.commission_rate),
                "image_url": p.image_url,
                "desi": str(p.desi),
                "default_carrier": p.default_carrier,
                "brand": p.brand,
                "return_rate": str(p.return_rate),
                "fast_delivery": p.fast_delivery,
                "current_stock": p.current_stock,
                "is_active": p.is_active,
                "variants": [
                    {
                        "id": v.id,
                        "title": v.title,
                        "barcode": v.barcode,
                        "marketplace_sku": v.marketplace_sku,
                        "cost_price": str(v.cost_price),
                        "cost_vat_rate": str(v.cost_vat_rate),
                        "desi": str(v.desi) if v.desi is not None else None,
                        "color": v.color,
                        "size": v.size,
                        "stock": v.stock,
                        "extra_cost_rate": str(v.extra_cost_rate),
                        "extra_cost_amount": str(v.extra_cost_amount),
                    } for v in p.variants.all()
                ]
            }
            data.append(product_data)

        # Total count should be count of distinct SKUs
        return paginator.get_paginated_response(data)

    def patch(self, request):
        """
        Updates product fields like desi and default_carrier, or variant fields like cost_price.
        Expects payload: 
        { "id": 123, "desi": "1.50", "default_carrier": "Aras Kargo" } -> Product update
        { "variant_id": 456, "cost_price": "250.00", "cost_vat_rate": "10", "desi": "0.5" } -> Variant update
        """
        user = request.user
        from core.models import UserProfile, Organization, Product, ProductVariant
        profile, _ = UserProfile.objects.get_or_create(user=user)
        org = profile.organization
        
        variant_id = request.data.get("variant_id")
        
        if variant_id:
            try:
                variant = ProductVariant.objects.get(id=variant_id, product__organization=org)
            except ProductVariant.DoesNotExist:
                return Response({"error": "Variant not found"}, status=404)
                
            if "cost_price" in request.data:
                variant.cost_price = Decimal(str(request.data["cost_price"]))
            if "cost_vat_rate" in request.data:
                variant.cost_vat_rate = Decimal(str(request.data["cost_vat_rate"]))
            if "desi" in request.data:
                if request.data["desi"] is None or request.data["desi"] == "":
                    variant.desi = None
                else:
                    variant.desi = Decimal(str(request.data["desi"]))
                    
            variant.save()
            return Response({
                "message": "Varyant güncellendi.",
                "variant_id": variant.id,
                "cost_price": str(variant.cost_price),
                "cost_vat_rate": str(variant.cost_vat_rate),
                "desi": str(variant.desi) if variant.desi is not None else None
            })
            
        else:
            product_id = request.data.get("id")
            if not product_id:
                 return Response({"error": "Product ID or Variant ID is required"}, status=400)
                 
            try:
                product = Product.objects.get(id=product_id, organization=org)
            except Product.DoesNotExist:
                return Response({"error": "Product not found"}, status=404)
                
            if "desi" in request.data:
                try:
                    product.desi = Decimal(str(request.data["desi"]))
                except Exception:
                    return Response({"error": "Invalid desi format"}, status=400)
                    
            if "default_carrier" in request.data:
                product.default_carrier = str(request.data["default_carrier"])
                
            if "return_rate" in request.data:
                try:
                    product.return_rate = Decimal(str(request.data["return_rate"]))
                except Exception:
                    return Response({"error": "Invalid return_rate format"}, status=400)

            if "vat_rate" in request.data:
                try:
                    product.vat_rate = Decimal(str(request.data["vat_rate"]))
                except Exception:
                    return Response({"error": "Invalid vat_rate format"}, status=400)
                    
            if "fast_delivery" in request.data:
                product.fast_delivery = bool(request.data["fast_delivery"])
                
            product.save()
            
            return Response({
                "id": product.id,
                "desi": str(product.desi),
                "default_carrier": product.default_carrier,
                "return_rate": str(product.return_rate),
                "vat_rate": str(product.vat_rate),
                "fast_delivery": product.fast_delivery,
                "message": "Ürün güncellendi."
            })

class OrderListView(APIView):
    """
    GET /api/orders/
    Sipariş listesini ve her siparişin detaylı kârlılık dökümünü döner.
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request):
        try:
            profile = request.user.profile
            org = profile.organization
        except Exception:
            return Response({"error": "Organizasyon bulunamadı"}, status=400)

        from core.models import Order
        from core.services.profit_calculator import ProfitCalculator
        from decimal import Decimal
        
        # Sadece bu organizasyona ait olan siparişleri getir, iptal edilenleri hariç tut
        # prefetch_related ile db call optimizasyonu
        CANCELLED_STATUSES = ['Cancelled', 'Canceled', 'İptal', 'iptal', 'CANCELLED', 'CANCELED']
        orders = Order.objects.filter(marketplace_account__organization=org).exclude(
            status__in=CANCELLED_STATUSES
        ).prefetch_related('items__product_variant__product', 'items__transactions').order_by('-order_date')
        
        min_date_str = request.query_params.get("min_date")
        max_date_str = request.query_params.get("max_date")

        if min_date_str and max_date_str:
            from datetime import datetime
            try:
                min_date = timezone.make_aware(datetime.strptime(min_date_str, "%Y-%m-%d"))
                max_date = timezone.make_aware(datetime.strptime(max_date_str + " 23:59:59", "%Y-%m-%d %H:%M:%S"))
                orders = orders.filter(order_date__gte=min_date, order_date__lte=max_date)
            except ValueError:
                pass
        else:
            # Sadece son 50 siparişi gönderelim (filtre yoksa)
            orders = orders[:50]

        data = []
        for order in orders:
            # Sipariş bazlı kârlılık hesabı — kargo ve hizmet bedeli TEK kez uygulanır
            calc = ProfitCalculator.calculate_for_order(order)
            kd   = calc["kdv_detail"]

            # Per-item display data (komisyon oranı ve fiyat için)
            items_data = []
            for item in order.items.all():
                item_r = ProfitCalculator.calculate_for_order_item(item)
                bd_item = item_r["breakdown"]
                title          = "Bilinmeyen Ürün"
                barcode        = item.sku
                commission_rate = item.applied_commission_rate or Decimal("0.00")
                image_url      = ""
                if item.product_variant and item.product_variant.product:
                    title      = item.product_variant.product.title
                    barcode    = item.product_variant.barcode
                    image_url  = item.product_variant.product.image_url or ""
                    if commission_rate == Decimal("0.00"):
                        commission_rate = item.product_variant.product.commission_rate
                items_data.append({
                    "id":               item.id,
                    "title":            title,
                    "barcode":          barcode,
                    "quantity":         item.quantity,
                    "sale_price_gross": str(item.sale_price_gross),
                    "commission_rate":  str(commission_rate),
                    "commission_cost":  str(bd_item.get("COMMISSION", Decimal("0.00"))),
                    "image_url":        image_url,
                })

            data.append({
                "id":             order.id,
                "order_number":   order.order_number,
                "order_date":     format_date_tr(order.order_date),
                "status":         order.status,
                "is_micro_export": order.channel == 'micro_export',
                "total_gross":    str(calc["total_sale"]),
                "total_profit":   str(calc["net_profit"]),
                "profit_margin":  str(calc["profit_margin"]),
                "profit_on_cost": str(calc["profit_on_cost"]),
                "items":          items_data,
                "aggregated_breakdown": {
                    "product_cost":       str(calc["product_cost"]),
                    "extra_cost":         str(calc["extra_product_cost"]),
                    "commission":         str(calc["commission"]),
                    "shipping_fee":       str(calc["cargo"]),
                    "cargo_source":       calc.get("cargo_source", "estimated"),
                    "service_fee":        str(calc["service_fee"]),
                    "withholding":        str(calc["withholding"]),
                    "net_kdv":            str(kd["net_kdv"]),
                    "satis_kdv":          str(kd["satis_kdv"]),
                    "alis_kdv":           str(kd["alis_kdv"]),
                    "kargo_kdv":          str(kd["kargo_kdv"]),
                    "komisyon_kdv":       str(kd["komisyon_kdv"]),
                    "hizmet_bedeli_kdv":  str(kd["hizmet_bedeli_kdv"]),
                },
            })

        return Response({"ok": True, "data": data})


class OrderExcelExportView(APIView):
    """
    GET /api/orders/export-excel/
    Sipariş kârlılık verilerini Excel olarak dışa aktarır.
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request):
        try:
            profile = request.user.profile
            org = profile.organization
        except Exception:
            return Response({"error": "Organizasyon bulunamadı"}, status=400)

        from core.models import Order
        from core.services.profit_calculator import ProfitCalculator
        from decimal import Decimal
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from io import BytesIO
        from django.http import HttpResponse

        orders = Order.objects.filter(
            marketplace_account__organization=org
        ).prefetch_related(
            'items__product_variant__product',
            'items__transactions'
        ).order_by('-order_date')

        min_date_str = request.query_params.get("min_date")
        max_date_str = request.query_params.get("max_date")

        if min_date_str and max_date_str:
            from datetime import datetime
            try:
                min_date = timezone.make_aware(datetime.strptime(min_date_str, "%Y-%m-%d"))
                max_date = timezone.make_aware(datetime.strptime(max_date_str + " 23:59:59", "%Y-%m-%d %H:%M:%S"))
                orders = orders.filter(order_date__gte=min_date, order_date__lte=max_date)
            except ValueError:
                pass
        else:
            orders = orders[:200]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sipariş Kârlılık"

        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        profit_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        loss_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

        columns = [
            "Sipariş Numarası", "Sipariş Tarihi", "Durum",
            "Sipariş Tutarı (₺)", "Ürün Maliyeti (₺)", "Ek Ürün Maliyeti (₺)",
            "Kargo Ücreti (₺)", "Komisyon Tutarı (₺)", "Hizmet Bedeli (₺)",
            "Stopaj Kesintisi (₺)", "Net KDV (₺)", "Kâr Tutarı (₺)",
            "Kâr Oranı (%)", "Kâr Marjı (%)",
        ]

        ws.append(columns)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        for i, _ in enumerate(columns, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = 22

        for order in orders:
            total_gross = Decimal("0.00")
            total_profit = Decimal("0.00")
            bd = {k: Decimal("0.00") for k in ["product_cost", "extra_cost", "commission", "shipping_fee", "service_fee", "withholding", "net_kdv"]}

            for item in order.items.all():
                pi = ProfitCalculator.calculate_for_order_item(item)
                total_gross += pi["gross_revenue"]
                total_profit += pi["net_profit"]
                b = pi["breakdown"]
                bd["product_cost"] += b.get("PRODUCT_COST", Decimal("0.00"))
                bd["extra_cost"] += pi.get("extra_product_cost", Decimal("0.00"))
                bd["commission"] += b.get("COMMISSION", Decimal("0.00"))
                bd["shipping_fee"] += b.get("SHIPPING_FEE", Decimal("0.00"))
                bd["service_fee"] += b.get("SERVICE_FEE", Decimal("0.00"))
                bd["withholding"] += b.get("WITHHOLDING", Decimal("0.00"))
                bd["net_kdv"] += pi.get("kdv_detail", {}).get("net_kdv", Decimal("0.00"))

            profit_margin = round((total_profit / total_gross) * 100, 2) if total_gross > 0 else Decimal("0.00")
            profit_on_cost = round((total_profit / bd["product_cost"]) * 100, 2) if bd["product_cost"] > 0 else Decimal("0.00")
            base_cost = bd["product_cost"] - bd["extra_cost"]

            ws.append([
                order.order_number,
                format_date_tr(order.order_date),
                order.status,
                float(total_gross),
                float(base_cost),
                float(bd["extra_cost"]),
                float(bd["shipping_fee"]),
                float(bd["commission"]),
                float(bd["service_fee"]),
                float(bd["withholding"]),
                float(bd["net_kdv"]),
                float(total_profit),
                float(profit_on_cost),
                float(profit_margin),
            ])

            last_row = ws.max_row
            profit_cell = ws.cell(last_row, 12)
            if total_profit > 0:
                profit_cell.fill = profit_fill
            elif total_profit < 0:
                profit_cell.fill = loss_fill

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="Siparis_Karliligi.xlsx"'
        return response





class CategoryAnalysisView(APIView):
    """
    GET /api/reports/categories/
    Kategori bazlı kârlılık verilerini döndürür. min_date/max_date filtreleri destekler.

    Düzeltmeler (Apr 2026):
    - Sadece Delivered siparişler dahil edilir (Cancelled/Returned hariç)
    - Kargo sipariş bazlı hesaplanır (item bazlı değil) → çoklu item siparişlerde kargo 1× sayılır
    - Her item'ın satış payı oranında sipariş maliyetleri kategoriye dağıtılır
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request):
        try:
            profile = request.user.profile
            org = profile.organization
        except Exception as e:
            return Response({"error": f"Organizasyon bulunamadı: {str(e)}"}, status=400)

        from core.models import OrderItem, Order
        from core.services.profit_calculator import ProfitCalculator
        from decimal import Decimal, ROUND_HALF_UP
        from datetime import datetime, time

        Q2 = Decimal("0.01")

        # ── Sadece teslim edilmiş siparişler ──────────────────────────────
        # Cancelled/Returned siparişler gerçek satış değil → hariç
        orders = Order.objects.filter(
            marketplace_account__organization=org,
            status='Delivered',
        )

        # Tarih filtresi
        min_date_str = request.query_params.get("min_date")
        max_date_str = request.query_params.get("max_date")
        if min_date_str and max_date_str:
            try:
                min_date = timezone.make_aware(datetime.strptime(min_date_str, "%Y-%m-%d"))
                max_date = timezone.make_aware(datetime.combine(datetime.strptime(max_date_str, "%Y-%m-%d"), time.max))
                orders = orders.filter(order_date__gte=min_date, order_date__lte=max_date)
            except ValueError:
                pass

        # Prefetch items ile N+1 sorgudan kaçın
        orders = orders.prefetch_related(
            Prefetch(
                'items',
                queryset=OrderItem.objects.select_related(
                    'product_variant', 'product_variant__product'
                ).prefetch_related('transactions')
            )
        )

        category_map = {}

        for order in orders:
            valid_items = [
                i for i in order.items.all()
                if i.product_variant and i.product_variant.product
                and i.product_variant.cost_price and i.product_variant.cost_price > Decimal("0")
            ]
            if not valid_items:
                continue

            # ── Sipariş bazlı kârlılık (kargo 1× hesaplanır) ─────────────
            order_result = ProfitCalculator.calculate_for_order(order)
            total_order_sale = order_result["total_sale"]
            if total_order_sale <= Decimal("0"):
                continue

            for item in valid_items:
                category = item.product_variant.product.category_name or "Kategorisiz"
                item_sale = item.sale_price_net or Decimal("0")

                # Bu item'ın sipariş içindeki satış payı
                share = item_sale / total_order_sale

                if category not in category_map:
                    category_map[category] = {
                        "category": category,
                        "product_count": set(),
                        "total_sold_quantity": 0,
                        "total_sales_amount": Decimal("0"),
                        "total_profit": Decimal("0"),
                        "total_cost": Decimal("0"),
                        "total_commission": Decimal("0"),
                        "total_cargo": Decimal("0"),
                    }

                category_map[category]["product_count"].add(item.product_variant.product.id)
                category_map[category]["total_sold_quantity"] += item.quantity
                category_map[category]["total_sales_amount"] += item_sale

                # Sipariş bazlı maliyetleri satış payına göre bu kategoriye dağıt
                category_map[category]["total_profit"]     += order_result["net_profit"] * share
                category_map[category]["total_cost"]       += order_result["product_cost"] * share
                category_map[category]["total_commission"] += order_result["commission"] * share
                category_map[category]["total_cargo"]      += order_result["cargo"] * share

        data = []
        for cat, stats in category_map.items():
            total_sales  = stats["total_sales_amount"]
            total_profit = stats["total_profit"].quantize(Q2, ROUND_HALF_UP)

            profit_margin = Decimal("0")
            if total_sales > Decimal("0"):
                profit_margin = (total_profit / total_sales * Decimal("100")).quantize(Q2, ROUND_HALF_UP)

            data.append({
                "id":                  cat,
                "category":            stats["category"],
                "product_count":       len(stats["product_count"]),
                "total_sold_quantity": stats["total_sold_quantity"],
                "total_sales_amount":  str(total_sales.quantize(Q2, ROUND_HALF_UP)),
                "total_profit":        str(total_profit),
                "total_commission":    str(stats["total_commission"].quantize(Q2, ROUND_HALF_UP)),
                "total_cargo":         str(stats["total_cargo"].quantize(Q2, ROUND_HALF_UP)),
                "profit_margin":       str(profit_margin),
            })

        data.sort(key=lambda x: Decimal(x["total_sales_amount"]), reverse=True)
        return Response({"ok": True, "data": data})


class ReturnAnalysisView(APIView):
    """
    GET /api/reports/returns/
    Sipariş bazlı iade/iptal analizi. Kargo zararı ve net zarar.
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    # Kargo firması adını CarrierFlatRate'e eşler
    _flat_rate_cache: dict = {}

    def _get_cargo_flat_rate(self, cargo_provider_name: str) -> "Decimal":
        from decimal import Decimal
        from core.models import CarrierFlatRate
        DEFAULT = Decimal("135.32")
        if not cargo_provider_name:
            return DEFAULT
        key = cargo_provider_name.lower()
        if key in self._flat_rate_cache:
            return self._flat_rate_cache[key]
        for carrier in CarrierFlatRate.objects.all():
            if carrier.carrier_name.lower() in key:
                self._flat_rate_cache[key] = carrier.rate_kdv_dahil
                return carrier.rate_kdv_dahil
        self._flat_rate_cache[key] = DEFAULT
        return DEFAULT

    def get(self, request):
        try:
            org = request.user.profile.organization
        except Exception:
            return Response({"error": "Organizasyon bulunamadı"}, status=400)

        import zoneinfo
        from datetime import datetime as dt_cls, time as dt_time, timedelta
        from decimal import Decimal, ROUND_HALF_UP

        tz = zoneinfo.ZoneInfo("Europe/Istanbul")
        Q2 = Decimal("0.01")
        # Sadece gerçekten kargo hareketi olan statüler — iptal/oluşturuldu dahil değil
        CARGO_LOSS_STATUSES = ["Returned", "UnDelivered"]
        self._flat_rate_cache = {}

        # ── Tarih aralığı ──────────────────────────────────────────────
        min_date_str = request.query_params.get("min_date")
        max_date_str = request.query_params.get("max_date")
        if min_date_str and max_date_str:
            try:
                min_date = dt_cls.strptime(min_date_str, "%Y-%m-%d").replace(tzinfo=tz)
                max_date = dt_cls.combine(
                    dt_cls.strptime(max_date_str, "%Y-%m-%d").date(), dt_time.max
                ).replace(tzinfo=tz)
            except ValueError:
                min_date = timezone.now() - timedelta(days=30)
                max_date = timezone.now()
        else:
            min_date = timezone.now() - timedelta(days=30)
            max_date = timezone.now()

        all_qs = Order.objects.filter(
            organization=org,
            order_date__gte=min_date,
            order_date__lte=max_date,
        )

        # Sadece Returned ve UnDelivered — kargoya verilip geri dönen siparişler
        returned_qs = all_qs.filter(
            status__in=CARGO_LOSS_STATUSES
        ).prefetch_related(
            Prefetch('items', queryset=OrderItem.objects.select_related('product_variant__product')),
        ).order_by('-order_date')

        # ── Claims'den ek siparişler (order_number ile eşleştirilir) ──
        from core.models import ReturnClaim
        claim_order_numbers = set(
            ReturnClaim.objects.filter(
                organization=org,
                claim_date__gte=min_date,
                claim_date__lte=max_date,
            ).exclude(order_number="").values_list("order_number", flat=True)
        )
        # Claims'de olan ama zaten CARGO_LOSS_STATUSES'ta olmayan siparişleri ekle
        claim_orders_qs = all_qs.filter(
            order_number__in=claim_order_numbers
        ).exclude(
            status__in=CARGO_LOSS_STATUSES
        ).prefetch_related(
            Prefetch('items', queryset=OrderItem.objects.select_related('product_variant__product')),
        )

        # ── Toplam satış: sadece Delivered + Shipped ─────────────────
        from django.db.models import Sum as DjSum
        total_sales = OrderItem.objects.filter(
            order__in=all_qs.filter(status__in=SALE_STATUSES)
        ).aggregate(s=DjSum('sale_price_gross'))['s'] or Decimal("0")

        # ── Sipariş bazlı hesaplama ────────────────────────────────────
        sum_return_amount  = Decimal("0")
        sum_return_qty     = 0
        sum_outgoing_cargo = Decimal("0")
        sum_incoming_cargo = Decimal("0")
        order_rows = []

        # Tüm işlenecek siparişler: (order, kargo_zararı_var_mı)
        orders_to_process = (
            [(o, True) for o in returned_qs] +
            [(o, True) for o in claim_orders_qs]
        )

        for order, has_cargo_loss in orders_to_process:
            items = list(order.items.all())
            if not items:
                continue

            order_sale_price = Decimal("0")
            order_qty        = 0
            display_name     = ""
            display_barcode  = ""

            for item in items:
                order_sale_price += item.sale_price_gross
                order_qty        += item.quantity
                if not display_name and item.product_variant and item.product_variant.product:
                    display_name    = item.product_variant.product.title
                    display_barcode = item.product_variant.barcode or ""

            if has_cargo_loss:
                outgoing_cargo = self._get_cargo_flat_rate(order.cargo_provider_name)
                incoming_cargo = outgoing_cargo
            else:
                outgoing_cargo = Decimal("0")
                incoming_cargo = Decimal("0")

            total_cargo_loss = (outgoing_cargo + incoming_cargo).quantize(Q2, ROUND_HALF_UP)
            net_loss = total_cargo_loss

            sum_return_amount  += order_sale_price
            sum_return_qty     += order_qty
            sum_outgoing_cargo += outgoing_cargo
            sum_incoming_cargo += incoming_cargo

            order_date_str = order.order_date.astimezone(tz).strftime("%d.%m.%Y %H:%M")
            order_rows.append({
                "order_number":     order.order_number or order.marketplace_order_id or "",
                "date":             order_date_str,
                "status":           order.status,
                "product_name":     display_name,
                "barcode":          display_barcode,
                "quantity":         order_qty,
                "sale_price":       str(order_sale_price.quantize(Q2, ROUND_HALF_UP)),
                "outgoing_cargo":   str(outgoing_cargo.quantize(Q2, ROUND_HALF_UP)),
                "incoming_cargo":   str(incoming_cargo.quantize(Q2, ROUND_HALF_UP)),
                "total_cargo_loss": str(total_cargo_loss),
                "commission":       "0.00",
                "net_loss":         str(net_loss),
            })

        order_rows.sort(key=lambda r: r["date"], reverse=True)

        # ── Özet ──────────────────────────────────────────────────────
        total_cargo_loss_sum = (sum_outgoing_cargo + sum_incoming_cargo).quantize(Q2, ROUND_HALF_UP)

        loss_ratio = Decimal("0")
        if total_sales > Decimal("0"):
            loss_ratio = (total_cargo_loss_sum / total_sales * Decimal("100")).quantize(Q2, ROUND_HALF_UP)

        return Response({
            "data": {
                "summary": {
                    "total_return_count":   sum_return_qty,
                    "total_return_amount":  str(sum_return_amount.quantize(Q2, ROUND_HALF_UP)),
                    "total_outgoing_cargo": str(sum_outgoing_cargo.quantize(Q2, ROUND_HALF_UP)),
                    "total_incoming_cargo": str(sum_incoming_cargo.quantize(Q2, ROUND_HALF_UP)),
                    "total_cargo_loss":     str(total_cargo_loss_sum),
                    "total_sales":          str(total_sales.quantize(Q2, ROUND_HALF_UP)),
                    "return_loss_ratio":    str(loss_ratio),
                },
                "orders": order_rows,
            }
        })


class ReturnLossView(APIView):
    """
    GET /api/reports/return-loss/
    getClaims API'den gelen iade kayıtlarını döndürür.
    Sadece aktif iade statüsleri: Accepted, WaitingInAction, Unresolved.
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    ACTIVE_STATUSES = {"Accepted", "WaitingInAction", "Unresolved", "InProgress"}

    def get(self, request):
        import zoneinfo
        from datetime import datetime as dt_cls, time as dt_time, timedelta
        from decimal import Decimal, ROUND_HALF_UP
        from core.models import ReturnClaim, ReturnClaimItem

        try:
            org = request.user.profile.organization
        except Exception:
            return Response({"error": "Organizasyon bulunamadı"}, status=400)

        tz = zoneinfo.ZoneInfo("Europe/Istanbul")
        Q2 = Decimal("0.01")

        min_date_str = request.query_params.get("min_date")
        max_date_str = request.query_params.get("max_date")
        if min_date_str and max_date_str:
            try:
                min_date = dt_cls.strptime(min_date_str, "%Y-%m-%d").replace(tzinfo=tz)
                max_date = dt_cls.combine(
                    dt_cls.strptime(max_date_str, "%Y-%m-%d").date(), dt_time.max
                ).replace(tzinfo=tz)
            except ValueError:
                min_date = timezone.now() - timedelta(days=30)
                max_date = timezone.now()
        else:
            min_date = timezone.now() - timedelta(days=30)
            max_date = timezone.now()

        claims_qs = ReturnClaim.objects.filter(
            organization=org,
            claim_status__in=list(self.ACTIVE_STATUSES),
            claim_date__gte=min_date,
            claim_date__lte=max_date,
        ).prefetch_related("claim_items").order_by("-claim_date")

        total_outgoing = Decimal("0")
        total_incoming = Decimal("0")
        total_item_count = 0
        total_refund = Decimal("0")
        claim_rows = []

        for claim in claims_qs:
            items = list(claim.claim_items.all())

            claim_outgoing = Decimal("0")
            claim_incoming = Decimal("0")
            claim_qty = 0
            product_name = ""
            barcode = ""
            customer_reason = ""

            for ci in items:
                claim_outgoing = max(claim_outgoing, ci.outgoing_cargo_cost)
                claim_incoming = max(claim_incoming, ci.incoming_cargo_cost)
                claim_qty += ci.quantity
                if not product_name:
                    product_name = ci.product_name
                    barcode = ci.barcode
                if not customer_reason and ci.customer_reason:
                    customer_reason = ci.customer_reason

            # Fallback: if no items yet (not synced at item level)
            if not items:
                claim_outgoing = claim.cargo_cost or Decimal("135.32")
                claim_incoming = claim.cargo_cost or Decimal("135.32")
                claim_qty = 1
                customer_reason = claim.reason

            total_cargo_loss = (claim_outgoing + claim_incoming).quantize(Q2, ROUND_HALF_UP)

            claim_date_str = ""
            if claim.claim_date:
                claim_date_str = claim.claim_date.astimezone(tz).strftime("%d.%m.%Y %H:%M")

            order_date_str = ""
            if claim.order_date:
                order_date_str = claim.order_date.astimezone(tz).strftime("%d.%m.%Y %H:%M")

            total_outgoing += claim_outgoing
            total_incoming += claim_incoming
            total_item_count += max(claim_qty, 1)
            total_refund += claim.refund_amount

            claim_rows.append({
                "claim_id": claim.claim_id,
                "order_number": claim.order_number,
                "claim_date": claim_date_str,
                "order_date": order_date_str,
                "claim_status": claim.claim_status,
                "product_name": product_name,
                "barcode": barcode,
                "quantity": claim_qty,
                "refund_amount": str(claim.refund_amount.quantize(Q2, ROUND_HALF_UP)),
                "outgoing_cargo": str(claim_outgoing.quantize(Q2, ROUND_HALF_UP)),
                "incoming_cargo": str(claim_incoming.quantize(Q2, ROUND_HALF_UP)),
                "total_cargo_loss": str(total_cargo_loss),
                "cargo_provider": claim.cargo_provider,
                "customer_reason": customer_reason or claim.reason,
            })

        total_cargo_loss_sum = (total_outgoing + total_incoming).quantize(Q2, ROUND_HALF_UP)

        return Response({
            "data": {
                "summary": {
                    "total_claim_count": len(claim_rows),
                    "total_item_count": total_item_count,
                    "total_refund_amount": str(total_refund.quantize(Q2, ROUND_HALF_UP)),
                    "total_outgoing_cargo": str(total_outgoing.quantize(Q2, ROUND_HALF_UP)),
                    "total_incoming_cargo": str(total_incoming.quantize(Q2, ROUND_HALF_UP)),
                    "total_cargo_loss": str(total_cargo_loss_sum),
                },
                "claims": claim_rows,
            }
        })


class AdsAnalysisView(APIView):
    """
    GET /api/reports/ads/
    Reklam analizi: filtrelenen tarih aralığında reklam harcamalarını, satış ve kâr verilerini döndürür.
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request):
        try:
            profile = request.user.profile
            org = profile.organization
        except Exception as e:
            return Response({"error": f"Organizasyon bulunamadı: {str(e)}"}, status=400)

        from core.models import OrderItem, Order, AdExpense
        from core.services.profit_calculator import ProfitCalculator
        from decimal import Decimal
        from datetime import datetime, time, date as date_cls

        Q2 = Decimal("0.01")
        ROUND_HALF_UP = __import__("decimal").ROUND_HALF_UP

        # ── Tarih filtresi ─────────────────────────────────────────────
        min_date_str = request.query_params.get("min_date")
        max_date_str = request.query_params.get("max_date")

        orders = Order.objects.filter(organization=org)
        ad_qs  = AdExpense.objects.filter(organization=org)

        if min_date_str and max_date_str:
            try:
                min_dt = timezone.make_aware(datetime.strptime(min_date_str, "%Y-%m-%d"))
                max_dt = timezone.make_aware(datetime.combine(datetime.strptime(max_date_str, "%Y-%m-%d"), time.max))
                min_d  = datetime.strptime(min_date_str, "%Y-%m-%d").date()
                max_d  = datetime.strptime(max_date_str, "%Y-%m-%d").date()
                orders = orders.filter(order_date__gte=min_dt, order_date__lte=max_dt)
                ad_qs  = ad_qs.filter(transaction_date__gte=min_d, transaction_date__lte=max_d)
            except ValueError:
                pass

        # ── Reklam & influencer giderleri ──────────────────────────────
        total_advertising = Decimal("0")
        total_influencer  = Decimal("0")
        transactions_list = []

        for exp in ad_qs.order_by("-transaction_date"):
            if exp.expense_type == AdExpense.ExpenseType.INFLUENCER:
                total_influencer += exp.amount
            else:
                total_advertising += exp.amount
            transactions_list.append({
                "date":             str(exp.transaction_date),
                "type":             exp.get_expense_type_display(),
                "transaction_type": exp.transaction_type,
                "amount":           str(exp.amount.quantize(Q2, ROUND_HALF_UP)),
                "description":      exp.description,
            })

        # ── CHE DeductionInvoices → reklam kategorisi ─────────────────
        from core.models import CheTransaction
        che_qs = CheTransaction.objects.filter(
            organization=org,
            source=CheTransaction.SOURCE_OTHER,
        )
        if min_date_str and max_date_str:
            try:
                min_dt = timezone.make_aware(datetime.strptime(min_date_str, "%Y-%m-%d"))
                max_dt = timezone.make_aware(datetime.combine(datetime.strptime(max_date_str, "%Y-%m-%d"), time.max))
                che_qs = che_qs.filter(transaction_date__gte=min_dt, transaction_date__lte=max_dt)
            except ValueError:
                pass

        for che in che_qs.order_by("-transaction_date"):
            desc = (che.description or "").lower()
            t_type = (che.transaction_type or "").lower()
            amount = che.debt if che.debt > 0 else che.credit
            if amount <= 0:
                continue
            if any(k in desc or k in t_type for k in ["reklam", "sponsorlu", "advertisement", "sponsored"]):
                expense_type_label = "Trendyol Reklam"
                total_advertising += amount
            elif any(k in desc or k in t_type for k in ["influencer"]):
                expense_type_label = "İnfluencer"
                total_influencer += amount
            else:
                continue  # Kargo, Platform Hizmet vb. reklam sayılmaz
            transactions_list.append({
                "date":             che.transaction_date.strftime("%Y-%m-%d"),
                "type":             expense_type_label,
                "transaction_type": che.transaction_type,
                "amount":           str(amount.quantize(Q2, ROUND_HALF_UP)),
                "description":      che.description or "",
            })

        # Fallback: FinancialTransaction ADS_COST (hiç veri yoksa)
        if not transactions_list:
            from core.models import FinancialTransaction
            ads_txns = FinancialTransaction.objects.filter(
                organization=org,
                transaction_type=FinancialTransactionType.ADS_COST.value,
            )
            if min_date_str and max_date_str:
                try:
                    min_dt = timezone.make_aware(datetime.strptime(min_date_str, "%Y-%m-%d"))
                    max_dt = timezone.make_aware(datetime.combine(datetime.strptime(max_date_str, "%Y-%m-%d"), time.max))
                    ads_txns = ads_txns.filter(occurred_at__gte=min_dt, occurred_at__lte=max_dt)
                except ValueError:
                    pass
            for txn in ads_txns:
                total_advertising += abs(txn.amount)
                transactions_list.append({
                    "date":             txn.occurred_at.strftime("%Y-%m-%d"),
                    "type":             "Trendyol Reklam",
                    "transaction_type": "ADS_COST",
                    "amount":           str(abs(txn.amount).quantize(Q2, ROUND_HALF_UP)),
                    "description":      "",
                })

        total_expense = (total_advertising + total_influencer).quantize(Q2, ROUND_HALF_UP)
        total_advertising = total_advertising.quantize(Q2, ROUND_HALF_UP)
        total_influencer  = total_influencer.quantize(Q2, ROUND_HALF_UP)

        # ── Satış & kâr ────────────────────────────────────────────────
        items = OrderItem.objects.filter(order__in=orders).select_related(
            'product_variant', 'product_variant__product', 'order'
        ).prefetch_related('transactions')

        total_sales  = Decimal("0.00")
        total_profit = Decimal("0.00")
        for item in items:
            profit_info = ProfitCalculator.calculate_for_order_item(item)
            total_sales  += profit_info["gross_revenue"]
            total_profit += profit_info["net_profit"]

        total_sales  = total_sales.quantize(Q2, ROUND_HALF_UP)
        total_profit = total_profit.quantize(Q2, ROUND_HALF_UP)

        # ── Oranlar ───────────────────────────────────────────────────
        advertising_sales_ratio  = Decimal("0.00")
        advertising_profit_ratio = Decimal("0.00")
        if total_sales > Decimal("0.00"):
            advertising_sales_ratio = (total_expense / total_sales * Decimal("100")).quantize(Q2, ROUND_HALF_UP)
        if total_profit > Decimal("0.00"):
            advertising_profit_ratio = (total_expense / total_profit * Decimal("100")).quantize(Q2, ROUND_HALF_UP)

        return Response({
            "data": {
                "summary": {
                    "total_advertising":        str(total_advertising),
                    "total_influencer":         str(total_influencer),
                    "total_expense":            str(total_expense),
                    "total_sales":              str(total_sales),
                    "total_profit":             str(total_profit),
                    "advertising_sales_ratio":  str(advertising_sales_ratio),
                    "advertising_profit_ratio": str(advertising_profit_ratio),
                },
                "transactions": transactions_list,
            }
        })


class PayoutsView(APIView):
    """
    GET /api/reports/payouts/
    Hakediş kontrol: PaymentOrder CHE işlemlerini döndürür.
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request):
        try:
            profile = request.user.profile
            org = profile.organization
        except Exception as e:
            return Response({"error": f"Organizasyon bulunamadı: {str(e)}"}, status=400)

        from core.models import CheTransaction
        from decimal import Decimal
        from datetime import datetime, time as dt_time

        Q2 = Decimal("0.01")
        ROUND_HALF_UP = __import__("decimal").ROUND_HALF_UP

        min_date_str = request.query_params.get("start_date") or request.query_params.get("min_date")
        max_date_str = request.query_params.get("end_date") or request.query_params.get("max_date")

        qs = CheTransaction.objects.filter(
            organization=org,
            source=CheTransaction.SOURCE_OTHER,
            transaction_type__icontains="Ödeme",
        )

        if min_date_str and max_date_str:
            try:
                min_dt = timezone.make_aware(datetime.strptime(min_date_str, "%Y-%m-%d"))
                max_dt = timezone.make_aware(datetime.combine(datetime.strptime(max_date_str, "%Y-%m-%d"), dt_time.max))
                qs = qs.filter(transaction_date__gte=min_dt, transaction_date__lte=max_dt)
            except ValueError:
                pass

        payments = []
        total_paid = Decimal("0")
        for che in qs.order_by("-transaction_date"):
            amount = che.debt if che.debt > 0 else che.credit
            total_paid += amount
            payments.append({
                "payment_order_id": che.payment_order_id,
                "payment_date":     format_date_tr(che.payment_date or che.transaction_date),
                "amount":           float(amount.quantize(Q2, ROUND_HALF_UP)),
                "description":      che.description or "",
                "status":           "Ödendi",
            })

        return Response({
            "data": {
                "summary": {
                    "total_paid":      float(total_paid.quantize(Q2, ROUND_HALF_UP)),
                    "total_pending":   0.0,
                    "payment_count":   len(payments),
                },
                "payments": payments,
            }
        })


import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

class ProductExcelExportView(APIView):
    """
    GET /api/products/export-excel/
    13 sütunlu Excel dosyası oluşturur.
    Sütunlar: Barkod, Model Kodu, Stok Kodu, Kategori İsmi, Ürün Adı,
              Trendyol Satış Fiyatı, Stok, Ürün Maliyeti (KDV Dahil),
              Maliyet KDV Oranı, Para Birimi, Ürün Desisi,
              Ekstra Maliyet (%), Ekstra Maliyet (TL)
    Mavi = Trendyol'dan otomatik doldurulur (salt okunur)
    Sarı  = Kullanıcı tarafından düzenlenir
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request):
        from core.models import UserProfile, Product
        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        org = profile.organization

        products = Product.objects.filter(organization=org, is_active=True).prefetch_related('variants').order_by(
            django_models.F('trendyol_created_at').desc(nulls_last=True)
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Urun Maliyetleri"

        # Column definitions: (header, editable)
        COLUMNS = [
            ("Barkod",                    False),  # 1  — Trendyol
            ("Model Kodu",                False),  # 2  — Trendyol
            ("Stok Kodu",                 False),  # 3  — Trendyol
            ("Kategori İsmi",             False),  # 4  — Trendyol
            ("Ürün Adı",                  False),  # 5  — Trendyol
            ("Trendyol Satış Fiyatı",     False),  # 6  — Trendyol
            ("Stok",                      False),  # 7  — Trendyol
            ("Ürün Maliyeti (KDV Dahil)", True),   # 8  — Kullanıcı
            ("Maliyet KDV Oranı",         True),   # 9  — Kullanıcı (default: Trendyol KDV)
            ("Para Birimi",               False),  # 10 — Trendyol
            ("Ürün Desisi",               True),   # 11 — Kullanıcı (default: 1)
            ("Ekstra Maliyet (%)",        True),   # 12 — Kullanıcı
            ("Ekstra Maliyet (TL)",       True),   # 13 — Kullanıcı
        ]

        info_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # mavi
        edit_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")  # sarı
        header_font = Font(bold=True)

        for col_num, (header_title, editable) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num, value=header_title)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.fill = edit_fill if editable else info_fill

        row_num = 2
        for product in products:
            for variant in product.variants.all():
                desi_val = float(variant.desi) if variant.desi is not None else float(product.desi if product.desi else 1)
                # Maliyet KDV Oranı: Kullanıcı girdiyse onu göster, yoksa Trendyol satış KDV'sini default al
                cost_vat = float(variant.cost_vat_rate) if variant.cost_vat_rate else float(product.vat_rate)

                ws.cell(row=row_num, column=1,  value=variant.barcode)
                ws.cell(row=row_num, column=2,  value=product.marketplace_sku)
                ws.cell(row=row_num, column=3,  value=variant.marketplace_sku or "")
                ws.cell(row=row_num, column=4,  value=product.category_name or "")
                ws.cell(row=row_num, column=5,  value=variant.title or product.title)
                ws.cell(row=row_num, column=6,  value=float(product.sale_price))
                ws.cell(row=row_num, column=7,  value=variant.stock)
                ws.cell(row=row_num, column=8,  value=float(variant.cost_price))
                ws.cell(row=row_num, column=9,  value=cost_vat)
                ws.cell(row=row_num, column=10, value=product.currency or "TRY")
                ws.cell(row=row_num, column=11, value=desi_val)
                ws.cell(row=row_num, column=12, value=float(variant.extra_cost_rate))
                ws.cell(row=row_num, column=13, value=float(variant.extra_cost_amount))
                row_num += 1

        # Auto column widths
        for col in ws.columns:
            max_length = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = "attachment; filename*=UTF-8''Urun_Maliyetleri.xlsx"
        wb.save(response)
        return response


class ProductExcelImportView(APIView):
    """
    POST /api/products/import-excel/
    Excel dosyasındaki maliyet, KDV, desi ve ekstra maliyet bilgilerini veritabanına kaydeder.
    Zorunlu: Barkod
    Desteklenen sütunlar: Ürün Maliyeti (KDV Dahil), Maliyet KDV Oranı, Ürün Desisi,
                          Ekstra Maliyet (%), Ekstra Maliyet (TL)
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def post(self, request):
        from core.models import UserProfile, ProductVariant
        from decimal import Decimal

        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        org = profile.organization

        file = request.FILES.get('file')
        if not file:
            return Response({"error": "Dosya bulunamadı."}, status=400)
        if not (file.name.endswith('.xlsx') or file.name.endswith('.xls')):
            return Response({"error": "Geçersiz dosya formatı. Lütfen .xlsx veya .xls yükleyin."}, status=400)

        def parse_decimal(val):
            if val is None or str(val).strip() == "":
                return None
            try:
                s = str(val).strip()
                if ',' in s and '.' in s:
                    s = s.replace('.', '').replace(',', '.')
                elif ',' in s:
                    s = s.replace(',', '.')
                return Decimal(s)
            except Exception:
                return None

        def parse_barcode(raw):
            if raw is None:
                return None
            if isinstance(raw, float):
                return str(int(raw)).strip()
            s = str(raw).strip()
            return s[:-2] if s.endswith('.0') else s

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active

            headers = {}
            for idx, cell in enumerate(ws[1]):
                if cell.value:
                    headers[str(cell.value).strip()] = idx

            def get_col(names):
                for n in names:
                    if n in headers:
                        return headers[n]
                return None

            idx_barkod       = get_col(["Barkod"])
            idx_cost         = get_col(["Ürün Maliyeti (KDV Dahil)", "KDV Dahil Maliyet", "Ürün Maliyeti ( KDV Dahil)"])
            idx_vat          = get_col(["Maliyet KDV Oranı", "KDV Oranı"])
            idx_desi         = get_col(["Ürün Desisi", "Desi"])
            idx_extra_rate   = get_col(["Ekstra Maliyet (%)"])
            idx_extra_amount = get_col(["Ekstra Maliyet (TL)"])

            if idx_barkod is None:
                return Response({"error": "Eksik sütun: 'Barkod' sütunu bulunamadı."}, status=400)
            if idx_cost is None:
                return Response({"error": "Eksik sütun: 'Ürün Maliyeti (KDV Dahil)' sütunu bulunamadı."}, status=400)

            updated_variants = 0
            unmatched_barcodes = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                barcode = parse_barcode(row[idx_barkod])
                if not barcode:
                    continue

                variant = ProductVariant.objects.select_related('product').filter(
                    barcode=barcode, product__organization=org
                ).first()
                if variant is None:
                    unmatched_barcodes.append(barcode)
                    continue

                try:
                    variant_fields = []
                    product_fields = []

                    clean_cost = parse_decimal(row[idx_cost])
                    if clean_cost is not None:
                        variant.cost_price = clean_cost
                        variant_fields.append("cost_price")

                    if idx_vat is not None:
                        clean_vat = parse_decimal(row[idx_vat])
                        if clean_vat is not None:
                            variant.cost_vat_rate = clean_vat
                            variant_fields.append("cost_vat_rate")

                    if idx_desi is not None:
                        clean_desi = parse_decimal(row[idx_desi])
                        if clean_desi is not None:
                            variant.desi = clean_desi
                            variant_fields.append("desi")
                            variant.product.desi = clean_desi
                            product_fields.append("desi")

                    if idx_extra_rate is not None:
                        clean_rate = parse_decimal(row[idx_extra_rate])
                        if clean_rate is not None:
                            variant.extra_cost_rate = clean_rate
                            variant_fields.append("extra_cost_rate")

                    if idx_extra_amount is not None:
                        clean_amount = parse_decimal(row[idx_extra_amount])
                        if clean_amount is not None:
                            variant.extra_cost_amount = clean_amount
                            variant_fields.append("extra_cost_amount")

                    if product_fields:
                        variant.product.save(update_fields=list(set(product_fields)))
                    if variant_fields:
                        variant.save(update_fields=variant_fields)
                    updated_variants += 1

                except Exception as e:
                    import logging
                    logging.getLogger(__name__).warning(f"Excel import row error (barcode={barcode}): {e}")

            msg = f"{updated_variants} varyant başarıyla güncellendi."
            if unmatched_barcodes:
                msg += f" {len(unmatched_barcodes)} barkod sistemde bulunamadı."

            return Response({"ok": True, "message": msg, "updated_count": updated_variants, "unmatched_count": len(unmatched_barcodes)})

        except Exception as e:
            return Response({"error": f"Excel dosyası okunamadı: {str(e)}"}, status=500)


class LivePerformanceView(APIView):
    """
    GET /api/live-performance/
    Canlı Performans sayfası için tek endpoint.
    Bugünün (veya filtre aralığının) sipariş kârlılığını, saatlik performansı,
    ürün heatmap'ini ve akıllı önerileri döndürür.
    Tüm finansal hesaplamalar backend'de Decimal ile yapılır.
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request):
        from core.models import UserProfile, Organization, ReturnClaim
        from core.services.profit_calculator import ProfitCalculator
        from datetime import datetime as dt_cls, time as dt_time, timedelta
        from django.db.models.functions import TruncHour
        from django.db.models import Sum as DjSum, Count
        from collections import defaultdict
        from rest_framework.pagination import PageNumberPagination

        user = request.user
        profile, _ = UserProfile.objects.get_or_create(user=user)
        org = profile.organization
        if not org:
            org, _ = Organization.objects.get_or_create(name=f"{user.username} Organizasyonu")
            profile.organization = org
            profile.save()

        # ── Date range (default: today) ──
        min_date_str = request.query_params.get("min_date")
        max_date_str = request.query_params.get("max_date")

        import zoneinfo
        tz_istanbul = zoneinfo.ZoneInfo("Europe/Istanbul")

        if min_date_str and max_date_str:
            try:
                min_date = dt_cls.strptime(min_date_str, "%Y-%m-%d").replace(tzinfo=tz_istanbul)
                max_date = dt_cls.combine(
                    dt_cls.strptime(max_date_str, "%Y-%m-%d").date(), dt_time.max
                ).replace(tzinfo=tz_istanbul)
            except ValueError:
                min_date = dt_cls.combine(timezone.localtime(timezone.now(), tz_istanbul).date(), dt_time.min).replace(tzinfo=tz_istanbul)
                max_date = dt_cls.combine(timezone.localtime(timezone.now(), tz_istanbul).date(), dt_time.max).replace(tzinfo=tz_istanbul)
        else:
            today = timezone.localtime(timezone.now(), tz_istanbul).date()
            min_date = dt_cls.combine(today, dt_time.min).replace(tzinfo=tz_istanbul)
            max_date = dt_cls.combine(today, dt_time.max).replace(tzinfo=tz_istanbul)

        # ── Base queryset ──
        orders_qs = Order.objects.filter(
            organization=org,
            order_date__gte=min_date,
            order_date__lte=max_date,
        ).exclude(
            status__in=["Cancelled", "Returned", "UnSupplied"]
        ).prefetch_related(
            'items__product_variant__product',
            'items__transactions'
        ).select_related('marketplace_account').order_by('-order_date')

        # ── Advanced Filters ──
        product_filter = request.query_params.get("product", "").strip()
        category_filter = request.query_params.get("category", "").strip()

        if product_filter:
            orders_qs = orders_qs.filter(
                items__product_variant__product__title__icontains=product_filter
            ).distinct()

        if category_filter:
            orders_qs = orders_qs.filter(
                items__product_variant__product__category_name__icontains=category_filter
            ).distinct()

        # ── Compute per-order profits ──
        all_order_rows = []
        product_agg = defaultdict(lambda: {
            "total_sales": 0, "total_revenue": Decimal("0"), "total_profit": Decimal("0"),
            "product_name": "", "return_rate": Decimal("0"), "category": ""
        })

        # Hourly aggregation buckets
        hourly_data = defaultdict(lambda: {"revenue": Decimal("0"), "profit": Decimal("0"), "order_count": 0})

        total_revenue = Decimal("0")
        total_profit = Decimal("0")
        total_commission = Decimal("0")
        total_cargo = Decimal("0")
        total_service_fee = Decimal("0")
        order_count = 0

        for order in orders_qs:
            calc = ProfitCalculator.calculate_for_order(order)
            order_profit = calc["net_profit"]
            order_sale = calc["total_sale"]
            order_commission = calc["commission"]
            order_cargo = calc["cargo"]
            order_service = calc["service_fee"]
            profit_margin = calc["profit_margin"]

            total_revenue += order_sale
            total_profit += order_profit
            total_commission += order_commission
            total_cargo += order_cargo
            total_service_fee += order_service
            order_count += 1

            # Determine main product name from first item
            product_name = "Bilinmeyen Ürün"
            product_cost = calc["product_cost"]
            return_rate = Decimal("0")
            category_name = ""

            first_item = order.items.all()[:1]
            if first_item:
                fi = first_item[0]
                if fi.product_variant and fi.product_variant.product:
                    product_name = fi.product_variant.product.title
                    return_rate = fi.product_variant.product.return_rate or Decimal("0")
                    category_name = fi.product_variant.product.category_name or ""

            # Return risk based on product return_rate
            if return_rate >= Decimal("20"):
                return_risk = "high"
            elif return_rate >= Decimal("10"):
                return_risk = "medium"
            else:
                return_risk = "low"

            kd = calc.get("kdv_detail", {})
            net_kdv = kd.get("net_kdv", Decimal("0"))

            order_row = {
                "order_number": order.order_number,
                "order_date": order.order_date.isoformat(),
                "product_name": product_name,
                "sale_price": str(order_sale),
                "cost": str(product_cost),
                "commission": str(order_commission),
                "cargo_cost": str(order_cargo),
                "tax": str(net_kdv),
                "net_profit": str(order_profit),
                "profit_margin": str(profit_margin),
                "return_risk": return_risk,
                "status": order.status,
                "cost_breakdown": {
                    "commission": str(order_commission),
                    "cargo": str(order_cargo),
                    "tax": str(net_kdv),
                    "service_fee": str(order_service),
                    "withholding": str(calc["withholding"]),
                    "product_cost": str(product_cost),
                }
            }
            all_order_rows.append(order_row)

            # Aggregate for product heatmap
            pk = product_name
            product_agg[pk]["product_name"] = product_name
            product_agg[pk]["total_sales"] += 1
            product_agg[pk]["total_revenue"] += order_sale
            product_agg[pk]["total_profit"] += order_profit
            product_agg[pk]["return_rate"] = return_rate
            product_agg[pk]["category"] = category_name

            # Hourly aggregation
            order_hour = order.order_date.replace(tzinfo=None)
            hour_key = order_hour.strftime("%H:00")
            hourly_data[hour_key]["revenue"] += order_sale
            hourly_data[hour_key]["profit"] += order_profit
            hourly_data[hour_key]["order_count"] += 1

        # ── Post-filter: profit range & loss-only ──
        min_profit = request.query_params.get("min_profit")
        max_profit = request.query_params.get("max_profit")
        loss_only = request.query_params.get("loss_only", "").lower() == "true"

        filtered_rows = all_order_rows
        if loss_only:
            filtered_rows = [r for r in filtered_rows if Decimal(r["net_profit"]) < Decimal("0")]
        if min_profit:
            try:
                mp = Decimal(min_profit)
                filtered_rows = [r for r in filtered_rows if Decimal(r["net_profit"]) >= mp]
            except Exception:
                pass
        if max_profit:
            try:
                mp = Decimal(max_profit)
                filtered_rows = [r for r in filtered_rows if Decimal(r["net_profit"]) <= mp]
            except Exception:
                pass

        # ── Sort support ──
        sort_by = request.query_params.get("sort_by", "order_date")
        sort_dir = request.query_params.get("sort_dir", "desc")
        reverse = sort_dir == "desc"

        sortable_decimal_fields = ["sale_price", "cost", "commission", "cargo_cost", "tax", "net_profit", "profit_margin"]
        if sort_by in sortable_decimal_fields:
            filtered_rows.sort(key=lambda r: Decimal(r.get(sort_by, "0")), reverse=reverse)
        elif sort_by == "order_date":
            filtered_rows.sort(key=lambda r: r.get("order_date", ""), reverse=reverse)
        elif sort_by == "product_name":
            filtered_rows.sort(key=lambda r: r.get("product_name", "").lower(), reverse=reverse)

        # ── Paginate orders ──
        total_count = len(filtered_rows)
        page = int(request.query_params.get("page", 1))
        page_size = int(request.query_params.get("page_size", 50))
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_rows = filtered_rows[start_idx:end_idx]

        # ── KPIs ──
        profit_margin_pct = Decimal("0")
        if total_revenue > Decimal("0"):
            profit_margin_pct = (total_profit / total_revenue * Decimal("100")).quantize(Decimal("0.01"))

        net_cash_inflow = total_revenue - total_commission - total_cargo - total_service_fee
        avg_order_profit = (total_profit / Decimal(str(order_count))).quantize(Decimal("0.01")) if order_count > 0 else Decimal("0")

        kpis = {
            "total_revenue": str(total_revenue.quantize(Decimal("0.01"))),
            "net_profit": str(total_profit.quantize(Decimal("0.01"))),
            "profit_margin": str(profit_margin_pct),
            "net_cash_inflow": str(net_cash_inflow.quantize(Decimal("0.01"))),
            "order_count": order_count,
            "avg_order_profit": str(avg_order_profit),
        }

        # ── Hourly performance (fill missing hours) ──
        hourly_list = []
        for h in range(24):
            hk = f"{h:02d}:00"
            entry = hourly_data.get(hk, {"revenue": Decimal("0"), "profit": Decimal("0"), "order_count": 0})
            hourly_list.append({
                "hour": hk,
                "revenue": str(entry["revenue"].quantize(Decimal("0.01"))),
                "profit": str(entry["profit"].quantize(Decimal("0.01"))),
                "order_count": entry["order_count"],
            })

        # ── Product heatmap (sorted by profit desc) ──
        heatmap = []
        for pk, agg in product_agg.items():
            margin = Decimal("0")
            if agg["total_revenue"] > Decimal("0"):
                margin = (agg["total_profit"] / agg["total_revenue"] * Decimal("100")).quantize(Decimal("0.01"))
            heatmap.append({
                "product_name": agg["product_name"],
                "category": agg["category"],
                "total_sales": agg["total_sales"],
                "total_revenue": str(agg["total_revenue"].quantize(Decimal("0.01"))),
                "total_profit": str(agg["total_profit"].quantize(Decimal("0.01"))),
                "margin": str(margin),
                "return_rate": str(agg["return_rate"]),
            })
        heatmap.sort(key=lambda x: Decimal(x["total_profit"]), reverse=True)

        # ── Smart Insights (rule-based) ──
        insights = []
        for pk, agg in product_agg.items():
            margin = Decimal("0")
            if agg["total_revenue"] > Decimal("0"):
                margin = (agg["total_profit"] / agg["total_revenue"] * Decimal("100")).quantize(Decimal("0.01"))

            # Negative profit warning
            if agg["total_profit"] < Decimal("0"):
                insights.append({
                    "type": "danger",
                    "title": "Zarar Eden Ürün",
                    "message": f"\"{agg['product_name']}\" bugün ₺{abs(agg['total_profit']).quantize(Decimal('0.01'))} zarar etti. Fiyatlandırma ve maliyetleri gözden geçirin.",
                    "product_name": agg["product_name"],
                })
            # High sales but low margin
            elif agg["total_sales"] >= 3 and margin < Decimal("10") and margin >= Decimal("0"):
                insights.append({
                    "type": "suggestion",
                    "title": "Fiyat Artışı Önerisi",
                    "message": f"\"{agg['product_name']}\" {agg['total_sales']} satış yaptı ama kâr marjı sadece %{margin}. Fiyat artışı düşünebilirsiniz.",
                    "product_name": agg["product_name"],
                })
            # High return rate
            if agg["return_rate"] >= Decimal("15"):
                insights.append({
                    "type": "warning",
                    "title": "Yüksek İade Riski",
                    "message": f"\"{agg['product_name']}\" ürününün iade oranı %{agg['return_rate']}. Ürün kalitesi veya açıklamalarını kontrol edin.",
                    "product_name": agg["product_name"],
                })

        return Response({
            "kpis": kpis,
            "orders": paginated_rows,
            "hourly_performance": hourly_list,
            "product_heatmap": heatmap,
            "insights": insights,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total_count": total_count,
            }
        })

class ProductProfitabilityView(APIView):
    """
    GET /api/reports/product-profitability/
    Barkod bazlı ürün kârlılık analizi.
    Her variant barcode için satış tutarı, kâr, iade kargo zararı hesaplar.
    Params: start_date, end_date, search, sort_by, sort_desc
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request):
        try:
            org = request.user.profile.organization
        except Exception:
            return Response({"error": "Organizasyon bulunamadı"}, status=400)

        if not org:
            return Response({"error": "Organizasyon bulunamadı"}, status=400)

        from datetime import datetime, time as dt_time
        import zoneinfo
        from collections import defaultdict

        tz = zoneinfo.ZoneInfo("Europe/Istanbul")

        start_date_str = request.query_params.get("start_date", "")
        end_date_str   = request.query_params.get("end_date", "")
        search         = request.query_params.get("search", "").strip().lower()
        sort_by        = request.query_params.get("sort_by", "total_sales")
        sort_desc      = request.query_params.get("sort_desc", "true").lower() != "false"

        # Date range
        if start_date_str and end_date_str:
            try:
                start_dt = datetime.strptime(start_date_str, "%Y-%m-%d").replace(tzinfo=tz)
                end_dt   = datetime.combine(
                    datetime.strptime(end_date_str, "%Y-%m-%d").date(), dt_time.max
                ).replace(tzinfo=tz)
            except ValueError:
                from datetime import timedelta
                end_dt   = timezone.now()
                start_dt = end_dt - timedelta(days=30)
        else:
            from datetime import timedelta
            end_dt   = timezone.now()
            start_dt = end_dt - timedelta(days=30)

        # Tüm siparişleri (iade dahil) çek — iade kargo zararı için gerekli
        orders_qs = Order.objects.filter(
            organization=org,
            order_date__gte=start_dt,
            order_date__lte=end_dt,
        ).prefetch_related(
            Prefetch('items', queryset=OrderItem.objects.select_related('order', 'product_variant__product').prefetch_related('transactions')),
        )

        # Satış sayılan: sadece Delivered + Shipped
        # Created/Picking/UnDelivered satış değil; Cancelled/UnSupplied kargo zararı yok
        SOLD_STATUSES     = set(SALE_STATUSES)
        RETURNED_STATUSES = set(RETURN_STATUSES + CANCEL_STATUSES)

        # barcode → aggregated data
        agg_map: dict = {}

        for order in orders_qs:
            is_sold     = order.status in SOLD_STATUSES
            is_returned = order.status in RETURNED_STATUSES

            if not is_sold and not is_returned:
                continue  # Shipped/Picking/Created: ne satış ne iade sayılır

            for item in order.items.all():
                variant = item.product_variant
                if not variant:
                    continue
                barcode = variant.barcode
                if not barcode:
                    continue
                product = variant.product
                if not product:
                    continue
                # Maliyetsiz ürünleri hesaplamadan çıkar
                if not variant.cost_price or variant.cost_price <= Decimal("0"):
                    continue

                # İlk görüldüğünde meta verileri kaydet
                if barcode not in agg_map:
                    agg_map[barcode] = {
                        "barcode":           barcode,
                        "product_name":      product.title,
                        "stock":             product.current_stock,
                        "model_code":        product.marketplace_sku or "",
                        "category":          product.category_name or "",
                        "total_sales":       Decimal("0.00"),
                        "total_profit":      Decimal("0.00"),
                        "total_cost":        Decimal("0.00"),
                        "return_cargo_loss": Decimal("0.00"),
                        "order_count":       0,
                        "return_count":      0,
                    }

                try:
                    profit_info = ProfitCalculator.calculate_for_order_item(item)
                except Exception:
                    continue

                bd       = profit_info.get("breakdown", {})
                shipping = bd.get(FinancialTransactionType.SHIPPING_FEE.value, Decimal("0.00"))
                cost     = bd.get(FinancialTransactionType.PRODUCT_COST.value, Decimal("0.00"))

                if is_returned:
                    # İade sipariş: kargo kaybını iade_kargo_zarari'ne ekle
                    agg_map[barcode]["return_cargo_loss"] += shipping
                    agg_map[barcode]["return_count"]      += item.quantity
                else:
                    # Delivered sipariş: satış adedi, tutar ve kâra ekle
                    agg_map[barcode]["total_sales"]  += profit_info["gross_revenue"]
                    agg_map[barcode]["total_profit"] += profit_info["net_profit"]
                    agg_map[barcode]["total_cost"]   += cost
                    agg_map[barcode]["order_count"]  += item.quantity  # Satış adedi = toplam adet

        # Search filtresi
        if search:
            agg_map = {
                k: v for k, v in agg_map.items()
                if search in v["barcode"].lower()
                or search in v["product_name"].lower()
                or search in v["model_code"].lower()
            }

        # Sonuçları oluştur
        results = []
        total_sales_sum  = Decimal("0.00")
        total_profit_sum = Decimal("0.00")
        total_rl_sum     = Decimal("0.00")

        q2 = Decimal("0.01")

        for barcode, agg in agg_map.items():
            satis  = agg["total_sales"]
            profit = agg["total_profit"]
            cost   = agg["total_cost"]

            # Kâr Oranı = Kâr / Satış × 100
            profit_rate = Decimal("0.00")
            if satis > Decimal("0.00"):
                profit_rate = (profit / satis * Decimal("100")).quantize(q2)

            # Kâr Marjı = Kâr / Maliyet × 100
            profit_margin = Decimal("0.00")
            if cost > Decimal("0.00"):
                profit_margin = (profit / cost * Decimal("100")).quantize(q2)

            total_sales_sum  += satis
            total_profit_sum += profit
            total_rl_sum     += agg["return_cargo_loss"]

            results.append({
                "barcode":           barcode,
                "product_name":      agg["product_name"],
                "stock":             agg["stock"],
                "model_code":        agg["model_code"],
                "category":          agg["category"],
                "total_sales":       str(satis.quantize(q2)),
                "total_profit":      str(profit.quantize(q2)),
                "return_cargo_loss": str(agg["return_cargo_loss"].quantize(q2)),
                "profit_rate":       str(profit_rate),
                "profit_margin":     str(profit_margin),
                "order_count":       agg["order_count"],
                "return_count":      agg["return_count"],
            })

        # Sıralama
        NUMERIC_SORT_KEYS = {
            "total_sales":       lambda x: Decimal(x["total_sales"]),
            "total_profit":      lambda x: Decimal(x["total_profit"]),
            "return_cargo_loss": lambda x: Decimal(x["return_cargo_loss"]),
            "profit_rate":       lambda x: Decimal(x["profit_rate"]),
            "profit_margin":     lambda x: Decimal(x["profit_margin"]),
            "stock":             lambda x: Decimal(str(x["stock"])),
        }
        key_fn = NUMERIC_SORT_KEYS.get(sort_by, lambda x: Decimal(x["total_sales"]))
        results.sort(key=key_fn, reverse=sort_desc)

        avg_profit_rate = Decimal("0.00")
        if total_sales_sum > Decimal("0.00"):
            avg_profit_rate = (total_profit_sum / total_sales_sum * Decimal("100")).quantize(q2)

        return Response({
            "results": results,
            "summary": {
                "total_sales":      str(total_sales_sum.quantize(q2)),
                "total_profit":     str(total_profit_sum.quantize(q2)),
                "total_return_loss": str(total_rl_sum.quantize(q2)),
                "avg_profit_rate":  str(avg_profit_rate),
            },
            "date_range": {
                "start_date": start_date_str or str(start_dt.date()),
                "end_date":   end_date_str   or str(end_dt.date()),
            }
        })


class ProductAnalysisView(APIView):
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request):
        try:
            org = request.user.profile.organization
        except Exception:
            return Response({"error": "Organizasyon bulunamadı"}, status=400)

        import zoneinfo
        from datetime import datetime as dt_cls, time as dt_time, timedelta
        from collections import defaultdict

        tz_istanbul = zoneinfo.ZoneInfo("Europe/Istanbul")

        # Accept both min_date/max_date and start_date/end_date for compatibility
        min_date_str = request.query_params.get("min_date") or request.query_params.get("start_date")
        max_date_str = request.query_params.get("max_date") or request.query_params.get("end_date")

        if min_date_str and max_date_str:
            try:
                min_date = dt_cls.strptime(min_date_str, "%Y-%m-%d").replace(tzinfo=tz_istanbul)
                max_date = dt_cls.combine(
                    dt_cls.strptime(max_date_str, "%Y-%m-%d").date(), dt_time.max
                ).replace(tzinfo=tz_istanbul)
            except ValueError:
                min_date = timezone.now() - timedelta(days=30)
                max_date = timezone.now()
        else:
            min_date = timezone.now() - timedelta(days=30)
            max_date = timezone.now()

        # Satış sayılan: sadece Delivered + Shipped
        SOLD_STATUSES     = set(SALE_STATUSES)
        RETURNED_STATUSES = set(RETURN_STATUSES + CANCEL_STATUSES)

        orders_qs = Order.objects.filter(
            organization=org,
            order_date__gte=min_date,
            order_date__lte=max_date,
        ).prefetch_related(
            'items__product_variant__product',
            'items__transactions',
        ).select_related('marketplace_account')

        product_agg = {}

        for order in orders_qs:
            is_sold     = order.status in SOLD_STATUSES
            is_returned = order.status in RETURNED_STATUSES
            if not is_sold and not is_returned:
                continue

            order_date_str = order.order_date.replace(tzinfo=None).strftime("%Y-%m-%d")

            for item in order.items.all():
                variant = item.product_variant
                if not variant or not variant.product:
                    continue
                p  = variant.product
                pk = p.barcode or p.marketplace_sku or p.title
                if not pk:
                    continue

                if pk not in product_agg:
                    product_agg[pk] = {
                        "barcode":            pk,
                        "title":              p.title,
                        "model_code":         variant.marketplace_sku or p.marketplace_sku or "",
                        "category":           p.category_name or "",
                        "stock":              p.current_stock,
                        "return_rate":        p.return_rate or Decimal("0"),
                        "total_sold_quantity": 0,
                        "revenue":            Decimal("0"),
                        "net_profit":         Decimal("0"),
                        "return_cargo_loss":  Decimal("0"),
                        "commission":         Decimal("0"),
                        "cargo":              Decimal("0"),
                        "tax":                Decimal("0"),
                        "service_fee":        Decimal("0"),
                        "cost":               Decimal("0"),
                        "trend":              defaultdict(lambda: {"revenue": Decimal("0"), "profit": Decimal("0")}),
                    }

                try:
                    calc = ProfitCalculator.calculate_for_order_item(item)
                except Exception:
                    continue

                bd = calc.get("breakdown", {})

                if is_returned:
                    cargo_loss = bd.get(FinancialTransactionType.SHIPPING_FEE.value, Decimal("0"))
                    product_agg[pk]["return_cargo_loss"] += cargo_loss
                else:
                    product_agg[pk]["total_sold_quantity"] += item.quantity
                    product_agg[pk]["revenue"]     += calc["gross_revenue"]
                    product_agg[pk]["net_profit"]  += calc["net_profit"]
                    product_agg[pk]["commission"]  += bd.get(FinancialTransactionType.COMMISSION.value, Decimal("0"))
                    product_agg[pk]["cargo"]       += bd.get(FinancialTransactionType.SHIPPING_FEE.value, Decimal("0"))
                    product_agg[pk]["service_fee"] += bd.get(FinancialTransactionType.SERVICE_FEE.value, Decimal("0"))
                    product_agg[pk]["tax"]         += calc.get("kdv_detail", {}).get("net_kdv", Decimal("0"))
                    product_agg[pk]["cost"]        += bd.get(FinancialTransactionType.PRODUCT_COST.value, Decimal("0"))
                    product_agg[pk]["trend"][order_date_str]["revenue"] += calc["gross_revenue"]
                    product_agg[pk]["trend"][order_date_str]["profit"]  += calc["net_profit"]

        max_sales = max([a["total_sold_quantity"] for a in product_agg.values()]) if product_agg else 1
        avg_sales = (sum([a["total_sold_quantity"] for a in product_agg.values()]) / len(product_agg)) if product_agg else 0

        results = []
        for pk, agg in product_agg.items():
            revenue = agg["revenue"]
            profit = agg["net_profit"]
            cost = agg["cost"]

            # Kâr Oranı = Kâr / Satış Tutarı × 100
            profit_rate = (profit / revenue * Decimal("100")) if revenue > 0 else Decimal("0")
            # Kâr Marjı = Kâr / Maliyet × 100
            profit_margin = (profit / cost * Decimal("100")) if cost > 0 else Decimal("0")

            return_rate_pct = agg["return_rate"]
            normalized_velocity = (Decimal(agg["total_sold_quantity"]) / Decimal(max_sales)) * Decimal("100")

            capped_margin = max(Decimal("0"), min(Decimal("100"), profit_rate))
            capped_return = max(Decimal("0"), min(Decimal("100"), return_rate_pct))

            score = (capped_margin * Decimal("0.4")) + (normalized_velocity * Decimal("0.3")) + ((Decimal("100") - capped_return) * Decimal("0.3"))
            score = max(Decimal("0"), min(score, Decimal("100"))).quantize(Decimal("0.1"))

            segment = "Standard"
            if profit > 0 and profit_rate > 15 and agg["total_sold_quantity"] > avg_sales:
                segment = "Cash Cow"
            elif profit > 0 and profit_rate <= 15 and agg["total_sold_quantity"] > avg_sales:
                segment = "Growth"
            elif profit <= 0 or agg["total_sold_quantity"] < (avg_sales * 0.2):
                segment = "Dead"

            tags = []
            if profit < 0: tags.append("Loss making")
            if 0 <= profit_rate < 10: tags.append("Low margin")
            if return_rate_pct > 10: tags.append("High return rate")
            if score >= 80: tags.append("Best performer")

            actions = []
            if segment == "Cash Cow" and agg["stock"] < 20: actions.append("Restock")
            if "Loss making" in tags or segment == "Growth": actions.append("Increase price")
            if segment == "Dead" and agg["stock"] > 50: actions.append("Stop selling")
            if score >= 70 and agg["stock"] > 10: actions.append("Increase ads")

            trend_arr = []
            for d in range(6, -1, -1):
                date_str = (timezone.localtime(timezone.now(), tz_istanbul) - timedelta(days=d)).strftime("%Y-%m-%d")
                t_data = agg["trend"].get(date_str, {"revenue": Decimal("0"), "profit": Decimal("0")})
                trend_arr.append({
                    "date": date_str,
                    "revenue": str(t_data["revenue"].quantize(Decimal("0.01"))),
                    "profit": str(t_data["profit"].quantize(Decimal("0.01")))
                })

            results.append({
                "barcode": agg["barcode"],
                "title": agg["title"],
                "model_code": agg["model_code"],
                "category": agg["category"],
                "stock": agg["stock"],
                "total_sold_quantity": agg["total_sold_quantity"],
                "total_sales_amount": str(revenue.quantize(Decimal("0.01"))),
                "total_profit": str(profit.quantize(Decimal("0.01"))),
                "return_cargo_loss": str(agg["return_cargo_loss"].quantize(Decimal("0.01"))),
                "profit_rate": str(profit_rate.quantize(Decimal("0.01"))),
                "profit_margin": str(profit_margin.quantize(Decimal("0.01"))),
                # Extra fields used by advanced analysis page
                "score": str(score),
                "segment": segment,
                "tags": tags,
                "actions": actions,
                "trend": trend_arr,
                "breakdown": {
                    "sale_price": str(revenue.quantize(Decimal("0.01"))),
                    "commission": str(agg["commission"].quantize(Decimal("0.01"))),
                    "cargo": str(agg["cargo"].quantize(Decimal("0.01"))),
                    "tax": str(agg["tax"].quantize(Decimal("0.01"))),
                    "return_loss": str(agg["return_cargo_loss"].quantize(Decimal("0.01"))),
                    "net_profit": str(profit.quantize(Decimal("0.01"))),
                }
            })

        results.sort(key=lambda x: Decimal(x["total_profit"]), reverse=True)
        return Response({"ok": True, "data": results})


class ProductProfitabilityExcelExportView(APIView):
    """
    GET /api/reports/product-profitability/export-excel/
    Ürün kârlılık verilerini Excel olarak dışa aktarır.
    """
    permission_classes = [IsAuthenticated, IsSubscribed]

    def get(self, request):
        try:
            org = request.user.profile.organization
        except Exception:
            return Response({"error": "Organizasyon bulunamadı"}, status=400)

        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from io import BytesIO
        from django.http import HttpResponse
        from datetime import datetime as dt_cls, time as dt_time, timedelta
        from collections import defaultdict

        tz_istanbul = timezone.get_fixed_timezone(180)
        min_date_str = request.query_params.get("min_date") or request.query_params.get("start_date")
        max_date_str = request.query_params.get("max_date") or request.query_params.get("end_date")

        if min_date_str and max_date_str:
            try:
                min_date = dt_cls.strptime(min_date_str, "%Y-%m-%d").replace(tzinfo=tz_istanbul)
                max_date = dt_cls.combine(
                    dt_cls.strptime(max_date_str, "%Y-%m-%d").date(), dt_time.max
                ).replace(tzinfo=tz_istanbul)
            except ValueError:
                min_date = timezone.localtime(timezone.now() - timedelta(days=30), tz_istanbul)
                max_date = timezone.localtime(timezone.now(), tz_istanbul)
        else:
            min_date = timezone.localtime(timezone.now() - timedelta(days=30), tz_istanbul)
            max_date = timezone.localtime(timezone.now(), tz_istanbul)

        orders_qs = Order.objects.filter(
            organization=org,
            order_date__gte=min_date,
            order_date__lte=max_date,
        ).exclude(
            status__in=["Cancelled", "Returned", "UnSupplied"]
        ).prefetch_related(
            'items__product_variant__product',
            'items__transactions'
        ).select_related('marketplace_account').order_by('order_date')

        returned_qs = Order.objects.filter(
            organization=org,
            order_date__gte=min_date,
            order_date__lte=max_date,
            status="Returned",
        ).prefetch_related(
            'items__product_variant__product',
            'items__transactions'
        ).select_related('marketplace_account')

        product_agg = defaultdict(lambda: {
            "barcode": "",
            "title": "",
            "model_code": "",
            "category": "",
            "stock": 0,
            "total_sold_quantity": 0,
            "revenue": Decimal("0"),
            "net_profit": Decimal("0"),
            "return_cargo_loss": Decimal("0"),
            "cost": Decimal("0"),
        })

        for order in orders_qs:
            calc = ProfitCalculator.calculate_for_order(order)
            fi = order.items.first()
            if not fi or not fi.product_variant or not fi.product_variant.product:
                continue

            p = fi.product_variant.product
            v = fi.product_variant
            pk = p.barcode or p.marketplace_sku or p.title

            agg = product_agg[pk]
            if not agg["barcode"]:
                agg["barcode"] = pk
                agg["title"] = p.title
                agg["model_code"] = v.marketplace_sku or p.marketplace_sku or ""
                agg["category"] = p.category_name
                agg["stock"] = p.current_stock

            agg["total_sold_quantity"] += 1
            agg["revenue"] += calc["total_sale"]
            agg["net_profit"] += calc["net_profit"]
            agg["cost"] += calc["product_cost"] + calc.get("extra_product_cost", Decimal("0"))

        for order in returned_qs:
            fi = order.items.first()
            if not fi or not fi.product_variant or not fi.product_variant.product:
                continue
            p = fi.product_variant.product
            pk = p.barcode or p.marketplace_sku or p.title
            if pk in product_agg:
                calc = ProfitCalculator.calculate_for_order(order)
                product_agg[pk]["return_cargo_loss"] += calc.get("cargo", Decimal("0"))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ürün Kârlılık"

        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        profit_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        loss_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

        columns = [
            "Barkod", "Ürün Adı", "Stok", "Model Kodu", "Kategori",
            "Satış Adedi", "Satış Tutarı (₺)", "Kâr Tutarı (₺)",
            "İade Kargo Zararı (₺)", "Kâr Oranı (%)", "Kâr Marjı (%)",
        ]
        ws.append(columns)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        col_widths = [18, 40, 8, 18, 20, 12, 18, 18, 20, 14, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        results = sorted(product_agg.values(), key=lambda x: x["net_profit"], reverse=True)
        for agg in results:
            revenue = agg["revenue"]
            profit = agg["net_profit"]
            cost = agg["cost"]
            profit_rate = float((profit / revenue * Decimal("100")).quantize(Decimal("0.01"))) if revenue > 0 else 0.0
            profit_margin = float((profit / cost * Decimal("100")).quantize(Decimal("0.01"))) if cost > 0 else 0.0

            row = [
                agg["barcode"],
                agg["title"],
                agg["stock"],
                agg["model_code"],
                agg["category"],
                agg["total_sold_quantity"],
                float(revenue.quantize(Decimal("0.01"))),
                float(profit.quantize(Decimal("0.01"))),
                float(agg["return_cargo_loss"].quantize(Decimal("0.01"))),
                profit_rate,
                profit_margin,
            ]
            ws.append(row)
            profit_cell = ws.cell(ws.max_row, 8)
            profit_cell.fill = profit_fill if profit >= 0 else loss_fill

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="Urun_Karliligi.xlsx"'
        return response


# ---------------------------------------------------------------------------
# Payment Views
# ---------------------------------------------------------------------------

class InitiatePaymentView(APIView):
    """POST /api/payments/initiate/ — PayTR ödeme token'ı oluştur."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from core.models import SubscriptionPlan, Payment
        from core.services.paytr_service import PayTRService

        plan_id = request.data.get("plan_id")
        if not plan_id:
            return Response({"error": "plan_id zorunludur."}, status=400)

        try:
            plan = SubscriptionPlan.objects.get(id=plan_id, is_active=True)
        except SubscriptionPlan.DoesNotExist:
            return Response({"error": "Plan bulunamadı."}, status=404)

        import time, random
        tmp_oid = f"TEMP_{request.user.id}_{int(time.time())}_{random.randint(100,999)}"
        payment = Payment.objects.create(
            user=request.user,
            plan=plan,
            amount=plan.price,
            status="pending",
            merchant_oid=tmp_oid,
        )

        user_ip = (
            request.META.get("HTTP_X_FORWARDED_FOR", "").split(",")[0].strip()
            or request.META.get("REMOTE_ADDR", "1.1.1.1")
        )

        paytr = PayTRService()
        try:
            merchant_oid, token = paytr.create_payment_token(
                payment=payment,
                user=request.user,
                user_ip=user_ip,
                callback_url=request.build_absolute_uri("/api/payments/callback/"),
                success_url=request.build_absolute_uri("/payment/success/"),
                fail_url=request.build_absolute_uri("/payment/fail/"),
            )
            payment.merchant_oid = merchant_oid
            payment.paytr_token = token
            payment.save()
            return Response({"token": token, "merchant_oid": merchant_oid})
        except Exception as e:
            payment.delete()
            return Response({"error": str(e)}, status=400)


class PayTRCallbackView(APIView):
    """POST /api/payments/callback/ — PayTR'den gelen ödeme sonucu."""
    permission_classes = []
    authentication_classes = []

    def post(self, request):
        from django.http import HttpResponse
        from core.models import Payment, UserSubscription
        from core.services.paytr_service import PayTRService

        paytr = PayTRService()
        post_data = request.POST.dict()

        if not paytr.verify_callback(post_data):
            return HttpResponse("PAYTR_ERROR")

        merchant_oid = post_data.get("merchant_oid", "")
        status = post_data.get("status", "")

        try:
            payment = Payment.objects.select_related("plan", "user").get(merchant_oid=merchant_oid)
            payment.paytr_response = post_data
            if status == "success":
                payment.status = "success"
                payment.save()
                # Aboneliği aktif et
                from django.utils import timezone
                from dateutil.relativedelta import relativedelta
                sub, _ = UserSubscription.objects.get_or_create(user=payment.user)
                sub.plan = payment.plan
                sub.status = "active"
                sub.admin_override = False
                if payment.plan and payment.plan.interval == "yearly":
                    sub.current_period_end = timezone.now() + relativedelta(years=1)
                else:
                    sub.current_period_end = timezone.now() + relativedelta(months=1)
                sub.save()
            else:
                payment.status = "failed"
                payment.save()
        except Payment.DoesNotExist:
            pass

        return HttpResponse("OK")


class PaymentHistoryView(APIView):
    """GET /api/payments/history/ — kullanıcının ödeme geçmişi."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from core.models import Payment
        payments = Payment.objects.filter(user=request.user).select_related("plan").order_by("-created_at")[:20]
        data = [
            {
                "id": p.id,
                "plan": p.plan.name if p.plan else "—",
                "amount": str(p.amount),
                "status": p.status,
                "merchant_oid": p.merchant_oid,
                "created_at": p.created_at.strftime("%d.%m.%Y %H:%M"),
            }
            for p in payments
        ]
        return Response(data)


class SubscriptionPlansView(APIView):
    """GET /api/subscription/plans/ — aktif abonelik planlarını listele."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from core.models import SubscriptionPlan
        plans = SubscriptionPlan.objects.filter(is_active=True).order_by("plan_tier", "interval")
        data = [
            {
                "id": p.id,
                "name": p.name,
                "price": str(p.price),
                "interval": p.interval,
                "plan_tier": p.plan_tier,
                "order_limit": p.order_limit,
                "store_limit": p.store_limit,
                "yearly_total": str(p.yearly_total) if p.yearly_total else None,
            }
            for p in plans
        ]
        return Response(data)
